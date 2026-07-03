"""
Evaluator: Orchestrates the complete evaluation pipeline for user simulation
"""
import json
import os
from typing import Dict, List, Tuple
from tqdm import tqdm
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

from config import *
from data_processor import load_user_data, filter_users_by_month
from prompt_builder import build_single_binary_prompt, build_single_continuous_prompt, build_single_text_prompt, get_binary_questions_for_action, get_all_questions_for_action, get_actual_used_history
from model_caller import ModelCaller, get_endpoints, assign_users_to_endpoints, stable_hash, DynamicTaskQueue
from metrics import calculate_all_binary_metrics, calculate_all_continuous_metrics, calculate_all_text_metrics, calculate_micro_macro_f1
from behavior_distribution import (
    analyze_prediction_distribution,
    generate_pr_curve_data,
    generate_distribution_comparison,
)
from datetime import datetime
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# 确保导入固定数据相关配置
try:
    from config import USE_FIXED_EXPERIMENT_DATA, FIXED_EXPERIMENT_DATA_PATH
except ImportError:
    USE_FIXED_EXPERIMENT_DATA = False
    FIXED_EXPERIMENT_DATA_PATH = "./dataset/experiment_data.json"

import numpy as np

from utils.prompts import prompt_for_evaluate_after_main
from z_5_practice_with_self_check import SelfPractice
self_practice = SelfPractice(user_id="", test_id="")
global OUTPUT_DIR
global RESULTS_DIR
global USE_EXP_MEM
global USE_EXP_MEM_PROGRESS
global EXP_MEM_DIR_PATH
global EXP_MEM_DIR_PATH_PROGRESS

def round_floats(obj, decimals=2):
    """递归地将字典/列表中的浮点数保留指定小数位，并处理 NaN/Inf 为 JSON 兼容格式"""
    if isinstance(obj, (float, np.floating)):
        if np.isnan(obj) or np.isinf(obj):
            return None  # JSON 不支持 NaN/Inf，转为 null
        return round(float(obj), decimals)
    elif isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, dict):
        return {k: round_floats(v, decimals) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [round_floats(item, decimals) for item in obj]
    else:
        return obj


def convert_metrics_to_percentage(metrics: Dict) -> Dict:
    """将指标中的百分比值（0-1范围）转换为百分数（0-100范围）
    
    处理以下类型的指标：
    - overall: F1, Precision, Recall, AUC, success_rate等
    - binary_metrics: 每个字段的F1, Precision, Recall, AUC等
    - text_metrics: BLEU, CharF1, BERTScore等
    - continuous_metrics: 保持不变（MAE, RMSE等不是百分比）
    
    注意：LogLoss 等损失函数类指标不应该被转换，它们的值范围不是 0-1 百分比
    """
    import copy
    metrics_copy = copy.deepcopy(metrics)
    
    # 需要转换的指标名称（包含这些关键词的指标会被乘以100）
    percentage_keywords = [
        'F1', 'Precision', 'Recall', 'AUC', 'Accuracy', 'ECE',
        'BLEU', 'CharF', 'BERTScore',
        'success_rate', 'positive_rate', 'Micro_', 'Macro_', 'R²'
    ]
    
    # 明确排除的指标（这些指标即使值在0-1范围内也不应该被转换）
    exclude_keywords = [
        'LogLoss', 'Loss', 'MAE', 'MSE', 'RMSE', 'mean_', 'count', 'TP', 'FP', 'FN', 'TN', 'relative_accuracy', 'symmetry_consistency'
    ]
    
    def should_convert(key: str) -> bool:
        """判断该指标是否应该转换为百分比"""
        # 首先检查是否在排除列表中
        if any(exclude in key for exclude in exclude_keywords):
            return False
        # 然后检查是否匹配转换关键词
        return any(keyword in key for keyword in percentage_keywords)
    
    def convert_dict(d: Dict) -> Dict:
        """递归转换字典中的百分比值"""
        result = {}
        for key, value in d.items():
            if isinstance(value, dict):
                result[key] = convert_dict(value)
            elif isinstance(value, (float, np.floating)) and should_convert(key):
                # 只转换 0-1 范围内的值（避免转换已经是百分比的值）
                if not np.isnan(value) and not np.isinf(value) and 0 <= value <= 1:
                    result[key] = float(value * 100)
                else:
                    result[key] = value
            else:
                result[key] = value
        return result
    
    # 转换各个部分
    if 'overall' in metrics_copy:
        metrics_copy['overall'] = convert_dict(metrics_copy['overall'])
    
    if 'binary_metrics' in metrics_copy:
        for field in metrics_copy['binary_metrics']:
            metrics_copy['binary_metrics'][field] = convert_dict(metrics_copy['binary_metrics'][field])
    
    if 'text_metrics' in metrics_copy:
        for field in metrics_copy['text_metrics']:
            metrics_copy['text_metrics'][field] = convert_dict(metrics_copy['text_metrics'][field])
    
    # continuous_metrics 也需要转换（R² 等是百分比，但 MAE, RMSE, MSE 等不是）
    if 'continuous_metrics' in metrics_copy:
        for field in metrics_copy['continuous_metrics']:
            metrics_copy['continuous_metrics'][field] = convert_dict(metrics_copy['continuous_metrics'][field])
    
    return metrics_copy


def generate_metrics_excel(json_data: Dict, output_file: str):
    """
    根据 metrics JSON 数据生成 Excel 表格
    
    Args:
        json_data: 包含 binary_metrics 的指标数据
        output_file: 输出的 Excel 文件路径
    """
    binary_metrics = json_data.get('binary_metrics', {})
    if not binary_metrics:
        return
    
    # 定义显示名称到实际键名的映射
    event_mapping = {
        'video': {
            'like': 'video_liked',
            'comment': 'video_commented',
            'share': 'video_shared',
            'collect': 'video_collected',
            'follow': 'video_followed',
            'complete': 'video_completed'
        },
        'live': {
            'like': 'live_liked',
            'comment': 'live_commented',
            'sent_gift': 'live_sent_gift',
            'follow': 'live_followed',
            'share': 'live_shared',
            'click_cart': 'live_clicked_cart'
        },
        'ad': {
            'like': 'ad_liked',
            'comment': 'ad_commented',
            'activate': 'ad_activated',
            'form_submit': 'ad_form_submitted'
        },
        'shop': {
            'add_to_cart': 'shop_added_to_cart',
            'order_success': 'shop_order_success'
        }
    }
    
    # 定义事件显示顺序（删除了 ad convert）
    display_order = {
        'video': ['like', 'comment', 'share', 'collect', 'follow', 'complete'],
        'live': ['like', 'comment', 'sent_gift', 'follow', 'share', 'click_cart'],
        'ad': ['like', 'comment', 'activate', 'form_submit'],
        'shop': ['add_to_cart', 'order_success']
    }
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "Metrics"
    
    # 写入第一行：类别标题（每个场景包含 avg 列）+ Overall 列
    row1_data = ['', '']  # Task 和 Model列
    for category in ['video', 'live', 'ad', 'shop']:
        events = display_order[category]
        # 事件数 + 1 (avg列)
        row1_data.extend([category] + [''] * len(events))
    # 添加 Overall 列（包含 Micro F1 和 Macro F1）
    row1_data.extend(['Overall', ''])
    
    for col, value in enumerate(row1_data, 1):
        ws.cell(1, col, value)
    
    # 写入第二行：事件名称 + avg + Overall 子列
    row2_data = ['Task', '']  # Model列留空
    for category in ['video', 'live', 'ad', 'shop']:
        for event in display_order[category]:
            row2_data.append(event)
        row2_data.append('avg')  # 每个场景后加 avg 列
    # 添加 Overall 子列
    row2_data.extend(['Micro', 'Macro'])
    
    for col, value in enumerate(row2_data, 1):
        ws.cell(2, col, value)
    
    # 获取 overall 数据（用于 Micro/Macro F1）
    overall = json_data.get('overall', {})
    
    # 写入指标行
    metrics_to_show = ['F1', 'AUC', 'LogLoss']
    col_idx_max = 3
    
    for row_offset, metric_name in enumerate(metrics_to_show, 3):
        ws.cell(row_offset, 1, metric_name)
        ws.cell(row_offset, 2, '')  # Model列
        
        col_idx = 3
        for category in ['video', 'live', 'ad', 'shop']:
            category_values = []  # 收集该场景下所有指标值，用于计算平均
            
            for event_display_name in display_order[category]:
                # 获取实际的键名
                full_event_name = event_mapping[category][event_display_name]
                
                cell_value = 0  # 默认值为 0
                if full_event_name in binary_metrics:
                    metric = binary_metrics[full_event_name]
                    
                    if metric_name == 'F1':
                        value = metric.get('F1', 0)
                        if value and value > 0:
                            ws.cell(row_offset, col_idx, value)
                            ws.cell(row_offset, col_idx).number_format = '0.00'
                            cell_value = value
                        else:
                            ws.cell(row_offset, col_idx, '')
                    elif metric_name == 'AUC':
                        value = metric.get('AUC', None)
                        if value is not None and str(value) != 'nan' and str(value) != 'None':
                            ws.cell(row_offset, col_idx, value)
                            ws.cell(row_offset, col_idx).number_format = '0.00'
                            cell_value = value
                        else:
                            ws.cell(row_offset, col_idx, 'nan')
                    elif metric_name == 'LogLoss':
                        value = metric.get('LogLoss', 0)
                        if value and value > 0:
                            ws.cell(row_offset, col_idx, value)
                            ws.cell(row_offset, col_idx).number_format = '0.00'
                            cell_value = value
                        else:
                            ws.cell(row_offset, col_idx, '')
                else:
                    ws.cell(row_offset, col_idx, '')
                
                category_values.append(cell_value)
                col_idx += 1
            
            # 写入该场景的平均值
            if category_values:
                avg_value = sum(category_values) / len(category_values)
                ws.cell(row_offset, col_idx, avg_value)
                ws.cell(row_offset, col_idx).number_format = '0.00'
            else:
                ws.cell(row_offset, col_idx, '')
            col_idx += 1
        
        # 写入 Overall 列（Micro F1 和 Macro F1）
        if metric_name == 'F1':
            # Micro F1
            micro_f1 = overall.get('Micro_F1', 0)
            if micro_f1 and micro_f1 > 0:
                ws.cell(row_offset, col_idx, micro_f1)
                ws.cell(row_offset, col_idx).number_format = '0.00'
            else:
                ws.cell(row_offset, col_idx, '')
            col_idx += 1
            
            # Macro F1
            macro_f1 = overall.get('Macro_F1', 0)
            if macro_f1 and macro_f1 > 0:
                ws.cell(row_offset, col_idx, macro_f1)
                ws.cell(row_offset, col_idx).number_format = '0.00'
            else:
                ws.cell(row_offset, col_idx, '')
            col_idx += 1
        else:
            # AUC 和 LogLoss 行在 Overall 列留空
            ws.cell(row_offset, col_idx, '')
            col_idx += 1
            ws.cell(row_offset, col_idx, '')
            col_idx += 1
            
        col_idx_max = col_idx
    
    # 合并类别标题单元格（包括 avg 列）
    col_idx = 3
    for category in ['video', 'live', 'ad', 'shop']:
        event_count = len(display_order[category]) + 1  # +1 for avg column
        if event_count > 1:
            ws.merge_cells(start_row=1, start_column=col_idx, 
                          end_row=1, end_column=col_idx + event_count - 1)
        col_idx += event_count
    
    # 合并 Overall 标题单元格（Micro + Macro = 2 列）
    ws.merge_cells(start_row=1, start_column=col_idx, 
                  end_row=1, end_column=col_idx + 1)
    
    # 设置样式（无边框）
    # 标题行
    for col in range(1, col_idx_max):
        ws.cell(1, col).font = Font(bold=True)
        ws.cell(1, col).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(2, col).font = Font(bold=True)
        ws.cell(2, col).alignment = Alignment(horizontal='center', vertical='center')
    
    # 指标名称列加粗
    for row in range(3, 6):
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')
    
    # 数据单元格居中
    for row in range(3, 6):
        for col in range(3, col_idx_max):
            ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    for col in range(3, col_idx_max):
        ws.column_dimensions[get_column_letter(col)].width = 11
    
    # 跟踪当前行位置
    current_row = 7  # 二分类指标后空一行
    
    # 添加 Micro/Macro F1 摘要
    overall = json_data.get('overall', {})
    if overall.get('Micro_F1') is not None or overall.get('Macro_F1') is not None:
        ws.cell(current_row, 1, "F1 汇总")
        ws.cell(current_row, 1).font = Font(bold=True)
        ws.merge_cells(start_row=current_row, start_column=1, 
                      end_row=current_row, end_column=4)
        current_row += 1
        
        # Micro F1
        ws.cell(current_row, 1, "Micro F1")
        ws.cell(current_row, 1).font = Font(bold=True)
        ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')
        micro_f1 = overall.get('Micro_F1', 0)
        if micro_f1:
            ws.cell(current_row, 2, micro_f1)
            ws.cell(current_row, 2).number_format = '0.0000'
        ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Macro F1
        ws.cell(current_row, 1, "Macro F1")
        ws.cell(current_row, 1).font = Font(bold=True)
        ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')
        macro_f1 = overall.get('Macro_F1', 0)
        if macro_f1:
            ws.cell(current_row, 2, macro_f1)
            ws.cell(current_row, 2).number_format = '0.0000'
        ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')
        current_row += 2  # 空一行
    
    # 添加连续值指标（如果有）
    continuous_metrics = json_data.get('continuous_metrics', {})
    if continuous_metrics:
        ws.cell(current_row, 1, "连续值预测指标")
        ws.cell(current_row, 1).font = Font(bold=True)
        ws.merge_cells(start_row=current_row, start_column=1, 
                      end_row=current_row, end_column=9)
        current_row += 1
        
        # 写入列标题（增加NMAE列）
        cont_headers = ["Field", "MAE", "NMAE(%)", "Relative Accuracy(%)", "Symm Consistency(%)", "MSE", "RMSE", "平均真实值", "样本数"]
        for col, header in enumerate(cont_headers, 1):
            ws.cell(current_row, col, header)
            ws.cell(current_row, col).font = Font(bold=True)
            ws.cell(current_row, col).alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # 字段名映射
        cont_field_display = {
            'video_watch_seconds': '视频观看时长',
            'live_watch_seconds': '直播观看时长',
            'ad_watch_seconds': '广告观看时长',
        }
        
        # 写入连续值指标数据
        for field, metrics in continuous_metrics.items():
            field_name = cont_field_display.get(field, field)
            ws.cell(current_row, 1, field_name)
            ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')
            
            # MAE
            mae = metrics.get('MAE', 0)
            if mae and not (isinstance(mae, float) and str(mae) == 'nan'):
                ws.cell(current_row, 2, mae)
                ws.cell(current_row, 2).number_format = '0.00'
            ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')
            
            # NMAE (归一化平均绝对误差，百分比形式)
            nmae = metrics.get('NMAE')
            if nmae is not None and not (isinstance(nmae, float) and str(nmae) == 'nan'):
                ws.cell(current_row, 3, nmae)
                ws.cell(current_row, 3).number_format = '0.00'
            else:
                ws.cell(current_row, 3, 'N/A')
            ws.cell(current_row, 3).alignment = Alignment(horizontal='center', vertical='center')
            
            # Relative Accuracy (相对准确率，百分比形式)
            rel_acc = metrics.get('relative_accuracy')
            if rel_acc is not None and not (isinstance(rel_acc, float) and str(rel_acc) == 'nan'):
                ws.cell(current_row, 4, rel_acc)
                ws.cell(current_row, 4).number_format = '0.00'
            else:
                ws.cell(current_row, 4, 'N/A')
            ws.cell(current_row, 4).alignment = Alignment(horizontal='center', vertical='center')
            
            # Symmetry Consistency (对称一致性，百分比形式)
            sym_cons = metrics.get('symmetry_consistency')
            if sym_cons is not None and not (isinstance(sym_cons, float) and str(sym_cons) == 'nan'):
                ws.cell(current_row, 5, sym_cons)
                ws.cell(current_row, 5).number_format = '0.00'
            else:
                ws.cell(current_row, 5, 'N/A')
            ws.cell(current_row, 5).alignment = Alignment(horizontal='center', vertical='center')
            
            # MSE
            mse = metrics.get('MSE', 0)
            if mse and not (isinstance(mse, float) and str(mse) == 'nan'):
                ws.cell(current_row, 6, mse)
                ws.cell(current_row, 6).number_format = '0.00'
            ws.cell(current_row, 6).alignment = Alignment(horizontal='center', vertical='center')
            
            # RMSE
            rmse = metrics.get('RMSE', 0)
            if rmse and not (isinstance(rmse, float) and str(rmse) == 'nan'):
                ws.cell(current_row, 7, rmse)
                ws.cell(current_row, 7).number_format = '0.00'
            ws.cell(current_row, 7).alignment = Alignment(horizontal='center', vertical='center')
            
            # 平均真实值
            mean_true = metrics.get('mean_true', 0)
            if mean_true and not (isinstance(mean_true, float) and str(mean_true) == 'nan'):
                ws.cell(current_row, 8, mean_true)
                ws.cell(current_row, 8).number_format = '0.00'
            ws.cell(current_row, 8).alignment = Alignment(horizontal='center', vertical='center')
            
            # 样本数
            sample_count = metrics.get('sample_count', 0)
            ws.cell(current_row, 9, sample_count)
            ws.cell(current_row, 9).alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
        
        current_row += 1  # 空一行
    
    # 添加文本预测指标（如果有）
    text_metrics = json_data.get('text_metrics', {})
    if text_metrics:
        ws.cell(current_row, 1, "文本预测指标")
        ws.cell(current_row, 1).font = Font(bold=True)
        ws.merge_cells(start_row=current_row, start_column=1, 
                      end_row=current_row, end_column=7)
        current_row += 1
        
        # 写入列标题（增加 BERTScore 列）
        ws.cell(current_row, 1, "Field")
        ws.cell(current_row, 2, "BLEU")
        ws.cell(current_row, 3, "CharF1")
        ws.cell(current_row, 4, "BERTScore_P")
        ws.cell(current_row, 5, "BERTScore_R")
        ws.cell(current_row, 6, "BERTScore_F1")
        ws.cell(current_row, 7, "样本数")
        
        for col in range(1, 8):
            ws.cell(current_row, col).font = Font(bold=True)
            ws.cell(current_row, col).alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # 写入文本指标数据
        for field, metrics in text_metrics.items():
            # 字段名映射
            field_display = {
                'search_keyword': '搜索关键词',
                'next_user_message': '用户回复',
            }.get(field, field)
            
            ws.cell(current_row, 1, field_display)
            ws.cell(current_row, 1).alignment = Alignment(horizontal='center', vertical='center')
            
            # BLEU
            bleu = metrics.get('BLEU', 0)
            if bleu and not (isinstance(bleu, float) and str(bleu) == 'nan'):
                ws.cell(current_row, 2, bleu)
                ws.cell(current_row, 2).number_format = '0.0000'
            else:
                ws.cell(current_row, 2, '')
            ws.cell(current_row, 2).alignment = Alignment(horizontal='center', vertical='center')
            
            # CharF1
            char_f1 = metrics.get('CharF1', 0)
            if char_f1 and not (isinstance(char_f1, float) and str(char_f1) == 'nan'):
                ws.cell(current_row, 3, char_f1)
                ws.cell(current_row, 3).number_format = '0.0000'
            else:
                ws.cell(current_row, 3, '')
            ws.cell(current_row, 3).alignment = Alignment(horizontal='center', vertical='center')
            
            # BERTScore_P
            bert_p = metrics.get('BERTScore_P', 0)
            if bert_p and not (isinstance(bert_p, float) and str(bert_p) == 'nan'):
                ws.cell(current_row, 4, bert_p)
                ws.cell(current_row, 4).number_format = '0.0000'
            else:
                ws.cell(current_row, 4, '')
            ws.cell(current_row, 4).alignment = Alignment(horizontal='center', vertical='center')
            
            # BERTScore_R
            bert_r = metrics.get('BERTScore_R', 0)
            if bert_r and not (isinstance(bert_r, float) and str(bert_r) == 'nan'):
                ws.cell(current_row, 5, bert_r)
                ws.cell(current_row, 5).number_format = '0.0000'
            else:
                ws.cell(current_row, 5, '')
            ws.cell(current_row, 5).alignment = Alignment(horizontal='center', vertical='center')
            
            # BERTScore_F1
            bert_f1 = metrics.get('BERTScore_F1', 0)
            if bert_f1 and not (isinstance(bert_f1, float) and str(bert_f1) == 'nan'):
                ws.cell(current_row, 6, bert_f1)
                ws.cell(current_row, 6).number_format = '0.0000'
            else:
                ws.cell(current_row, 6, '')
            ws.cell(current_row, 6).alignment = Alignment(horizontal='center', vertical='center')
            
            # 样本数
            sample_count = metrics.get('sample_count', 0)
            ws.cell(current_row, 7, sample_count)
            ws.cell(current_row, 7).alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
    
    wb.save(output_file)
    print(f"  - Excel: {output_file}")


class UserSimulationEvaluator:
    """用户模拟评估器"""
    
    def __init__(self, model_config: Dict, experiment_name: str = None, run_timestamp: str = None, run_date: str = None, max_history_tokens: int = None, max_history_days: int = None, history_scene_filter: str = None):
        """
        初始化评估器
        
        Args:
            model_config: 模型配置字典
            experiment_name: 实验名称
            run_timestamp: 运行时间戳（用于统一保存路径，精确到秒）
            run_date: 运行日期（用于日期目录）
            max_history_tokens: 历史行为最大 token 数限制，如果为 None 则使用 config.MAX_HISTORY_TOKENS
            max_history_days: 只保留近 N 天的历史行为，如果为 None 则不限制天数
            history_scene_filter: 历史行为场景过滤（如"视频浏览"、"直播间"等），为空则不过滤
        """
        self.model_config = model_config
        self.endpoints = get_endpoints(model_config)
        
        # 如果只有单端点，直接创建 model_caller
        if len(self.endpoints) <= 1:
            self.model_caller = ModelCaller(model_config=model_config)
        else:
            self.model_caller = None  # 多端点模式下在 evaluate_all 中按需创建
        
        self.results = []
        self.experiment_name = experiment_name or "default"
        self.model_name = model_config.get("name", "unknown")
        self.max_history_tokens = max_history_tokens
        self.max_history_days = max_history_days  # 只保留近 N 天的历史行为
        self.history_scene_filter = history_scene_filter  # 历史行为场景过滤
        
        # 使用上海时区
        shanghai_tz = pytz.timezone('Asia/Shanghai')
        now_shanghai = datetime.now(shanghai_tz)
        
        self.run_date = run_date or now_shanghai.strftime("%Y-%m-%d")
        self.run_timestamp = run_timestamp or now_shanghai.strftime("%Y-%m-%d_%H-%M-%S")
        
        # 打印端点信息
        if len(self.endpoints) > 1:
            total_workers = sum(ep["max_workers"] for ep in self.endpoints)
            print(f"📡 多端点模式，共 {len(self.endpoints)} 个端点，总并发 {total_workers}:")
            for i, ep in enumerate(self.endpoints):
                print(f"   [{i+1}] {ep['url']} (并发: {ep['max_workers']})")
        
    def evaluate_single_action(
        self,
        user_id: str,
        user_profile: str,
        action_history: List[Dict],
        test_action: Dict,
        model_caller: ModelCaller = None
    ) -> Dict:
        """
        评估单个行为预测（二分类模式：每个问题单独调用，只输出 Yes/No）
        
        Args:
            model_caller: 可选，指定使用的 model_caller（多端点模式）
        
        Returns:
            {
                "user_id": str,
                "action_type": str,
                "timestamp": str,
                "questions": List[Dict],  # 每个问题的真实值和预测值
                "prompts": List[str],  # 每个问题的 prompt
                "total_prompt_length": int,  # 所有 prompt 总字符数
                "model_responses": List[str],  # 每个问题的模型原始响应
                "total_response_length": int,  # 所有响应总字符数
                "success": bool,  # 是否成功获取预测（至少有一个成功）
                "filtered": bool,  # 是否被过滤（play_duration为0）
                "binary_mode": bool,  # 是否使用二分类模式
                "failed_questions": List[Dict],  # 失败的问题记录
            }
        """
        # 使用传入的 model_caller 或默认的
        caller = model_caller or self.model_caller
        # 获取该行为的所有预测问题（包括 binary + continuous）
        all_questions = get_all_questions_for_action(test_action)
        
        # 如果没有问题，检查是否应该被过滤
        from prompt_builder import should_filter_action
        if should_filter_action(test_action):
            return {
                "user_id": user_id,
                "action_type": test_action.get("type", "unknown"),
                "timestamp": test_action.get("timestamp", "unknown"),
                "scene_info": None,
                "prompts": [],
                "total_prompt_length": 0,
                "model_responses": [],
                "total_response_length": 0,
                "success": False,
                "filtered": True,
                "questions": [],
                "binary_mode": True,
                "failed_questions": [],
            }
        
        # 如果没有问题，返回空结果
        if not all_questions:
            return {
                "user_id": user_id,
                "action_type": test_action.get("type", "unknown"),
                "timestamp": test_action.get("timestamp", "unknown"),
                "scene_info": None,
                "prompts": [],
                "total_prompt_length": 0,
                "model_responses": [],
                "total_response_length": 0,
                "success": True,
                "filtered": False,
                "questions": [],
                "binary_mode": True,
                "failed_questions": [],
            }
        
        # 提取场景信息
        scene_info = {
            "type": test_action.get("type", "unknown"),
            "timestamp": test_action.get("timestamp", "unknown"),
            "context": test_action.get("context", {}),
            "action": test_action.get("action", []),
        }
        
        # 获取实际使用的历史行为统计信息
        # 必须传递 max_history_days 和 reference_timestamp 才能正确过滤天数
        history_info = get_actual_used_history(
            action_history, 
            max_history_tokens=self.max_history_tokens,
            max_history_days=self.max_history_days,
            reference_timestamp=test_action.get("timestamp")
        )
        actual_used_actions = history_info["actual_used_actions"]
        
        # 统计实际使用的历史行为场景分布
        history_scene_counts = {}
        history_action_counts = {}
        history_timestamps = []  # 收集所有历史行为的时间戳
        
        for action in actual_used_actions:
            scene_type = action.get("type", "Unknown")
            history_scene_counts[scene_type] = history_scene_counts.get(scene_type, 0) + 1
            
            # 收集时间戳
            timestamp = action.get("timestamp", "")
            if timestamp:
                history_timestamps.append(timestamp)
            
            # 统计场景内的行为
            actions_list = action.get("action", [])
            if isinstance(actions_list, list):
                for act in actions_list:
                    act_type = act.get("type", "unknown")
                    key = f"{scene_type}_{act_type}"
                    history_action_counts[key] = history_action_counts.get(key, 0) + 1
        
        # 计算历史行为的时间分布统计
        history_time_stats = self._calculate_history_time_stats(history_timestamps)
        
        result = {
            "user_id": user_id,
            "action_type": test_action.get("type", "unknown"),
            "timestamp": test_action.get("timestamp", "unknown"),
            "scene_info": scene_info,
            "prompts": [],
            "total_prompt_length": 0,
            "model_responses": [],
            "total_response_length": 0,
            "total_prompt_tokens": 0,  # API 返回的真实 prompt token 数
            "total_completion_tokens": 0,  # API 返回的真实 completion token 数
            "total_cached_tokens": 0,  # 缓存命中的 token 数
            "success": False,
            "filtered": False,
            "questions": [],
            "binary_mode": True,
            "failed_questions": [],
            # 实际使用的历史行为统计
            "history_stats": {
                "original_count": history_info["original_count"],
                "filtered_count": history_info["filtered_count"],
                "actual_used_count": history_info["actual_used_count"],
                "actual_used_tokens": history_info["actual_used_tokens"],
                "scene_distribution": history_scene_counts,
                "action_distribution": history_action_counts,
                # 历史行为时间分布统计
                "earliest_timestamp": history_time_stats.get("earliest_timestamp"),
                "latest_timestamp": history_time_stats.get("latest_timestamp"),
                "time_span_days": history_time_stats.get("time_span_days"),
                "avg_actions_per_day": history_time_stats.get("avg_actions_per_day"),
            },
        }
        
        # 对每个问题单独调用模型
        successful_predictions = 0
        
        exp_memory_prompt = ""
        if USE_EXP_MEM == 101:
            exp_memory_path = f"{EXP_MEM_DIR_PATH}/{test_action['user_id']}_0.json"
            exp_memory_progress_path = f"{EXP_MEM_DIR_PATH_PROGRESS}/{test_action['user_id']}_0.json"

            if os.path.exists(exp_memory_path):
                try:
                    if USE_EXP_MEM_PROGRESS != 999:
                        exp_memory = json.load(open(exp_memory_progress_path))[f"round_{USE_EXP_MEM_PROGRESS}"]['memory_snapshot_after']
                    else:
                        exp_memory = json.load(open(exp_memory_path))
                    exp_memory_str = self_practice.format_exp_memory(exp_memory, with_proposal=0)
                    exp_memory_prompt = prompt_for_evaluate_after_main.format(memory=exp_memory_str)
                except Exception as e:
                    print(f"[Warning] Failed to load exp_memory from {exp_memory_path}: {e}")
            else:
                print(f"\n[Warning] exp_memory file not found, skip: {exp_memory_path}")

        for question_info in all_questions:
            question_type = question_info.get("type", "binary")
            
            # 根据问题类型构建不同的 prompt
            if question_type == "binary":
                # 二分类问题使用 Yes/No prompt
                prompt_data = build_single_binary_prompt(
                    user_profile,
                    action_history,
                    test_action,
                    question_info,
                    max_history_tokens=self.max_history_tokens,
                    max_history_days=self.max_history_days
                )
            elif question_type == "continuous":
                # 连续值问题使用数字输出 prompt
                prompt_data = build_single_continuous_prompt(
                    user_profile,
                    action_history,
                    test_action,
                    question_info,
                    max_history_tokens=self.max_history_tokens,
                    max_history_days=self.max_history_days
                )
            elif question_type == "text":
                # 文本预测问题使用文本输出 prompt
                prompt_data = build_single_text_prompt(
                    user_profile,
                    action_history,
                    test_action,
                    question_info,
                    max_history_tokens=self.max_history_tokens,
                    max_history_days=self.max_history_days
                )
            else:
                # 未知类型，跳过
                continue
            
            if prompt_data is None:
                continue
            
            prompt = prompt_data["prompt"]

            if exp_memory_prompt != "":
                prompt = f"{prompt}\n\n{exp_memory_prompt}" 

            result["prompts"].append(prompt)
            result["total_prompt_length"] += len(prompt)
            
            # 根据问题类型调用不同的预测方法
            if question_type == "binary":
                # 二分类预测
                prediction_result = caller.call_binary_classification(prompt)
            elif question_type == "continuous":
                # 连续值预测
                prediction_result = caller.call_continuous_prediction(prompt)
            elif question_type == "text":
                # 文本预测
                prediction_result = caller.call_text_prediction(prompt)
            
            # 记录原始响应
            raw_output = prediction_result.get("raw_output", "")
            result["model_responses"].append(raw_output)
            result["total_response_length"] += len(raw_output)
            
            # 累加 API 返回的真实 token 使用量
            result["total_prompt_tokens"] += prediction_result.get("prompt_tokens", 0)
            result["total_completion_tokens"] += prediction_result.get("completion_tokens", 0)
            result["total_cached_tokens"] += prediction_result.get("cached_tokens", 0)
            
            # 记录问题结果
            question_result = {
                "type": question_info["type"],
                "field": question_info["field"],
                "true_value": question_info["true_value"],
                "predicted_value": prediction_result.get("prediction"),
                "predicted_label": prediction_result.get("predicted_label"),  # 基于YES/NO直接判定的标签
                "prediction_method": prediction_result.get("method"),
                "raw_output": raw_output,
                "retry_count": prediction_result.get("retry_count", 0),
                "logprob_yes": prediction_result.get("logprob_yes"),
                "logprob_no": prediction_result.get("logprob_no"),
                # 连续值预测的归一化因子（如视频时长），用于计算 NMAE
                "video_duration": question_info.get("video_duration"),
                # 每个 prompt 的 token 统计
                "prompt_tokens": prediction_result.get("prompt_tokens", 0),
                "completion_tokens": prediction_result.get("completion_tokens", 0),
                "cached_tokens": prediction_result.get("cached_tokens", 0),
                # 该 prompt 中使用的历史行为数量
                "history_action_count": history_info["actual_used_count"],
                "history_token_count": history_info["actual_used_tokens"],
                # 历史行为时间分布统计
                "history_earliest_time": history_time_stats.get("earliest_timestamp"),
                "history_latest_time": history_time_stats.get("latest_timestamp"),
                "history_time_span_days": history_time_stats.get("time_span_days"),
                "history_avg_actions_per_day": history_time_stats.get("avg_actions_per_day"),
                # 是否有之前的用户发言（用于控制 BERTScore 计算）
                "has_prior_user_speech": question_info.get("has_prior_user_speech"),
            }
            
            if prediction_result.get("success"):
                successful_predictions += 1
                result["questions"].append(question_result)
            else:
                # 记录失败的问题
                question_result["error"] = prediction_result.get("error")
                result["failed_questions"].append(question_result)
        
        # 如果至少有一个预测成功，则认为整体成功
        result["success"] = successful_predictions > 0
        
        return result
    
    def _calculate_history_time_stats(self, timestamps: List[str]) -> Dict:
        """
        计算历史行为的时间分布统计
        
        Args:
            timestamps: 历史行为时间戳列表
            
        Returns:
            {
                "earliest_timestamp": str,  # 最早的历史行为时间
                "latest_timestamp": str,    # 最后一个历史行为时间
                "time_span_days": float,    # 时间跨度（天）
                "avg_actions_per_day": float,  # 平均每天行为数
            }
        """
        if not timestamps:
            return {
                "earliest_timestamp": None,
                "latest_timestamp": None,
                "time_span_days": None,
                "avg_actions_per_day": None,
            }
        
        from datetime import datetime
        
        # 解析时间戳
        parsed_times = []
        for ts in timestamps:
            if not ts:
                continue
            try:
                # 尝试多种时间格式
                for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"]:
                    try:
                        parsed_times.append(datetime.strptime(ts, fmt))
                        break
                    except ValueError:
                        continue
            except Exception:
                continue
        
        if not parsed_times:
            return {
                "earliest_timestamp": timestamps[0] if timestamps else None,
                "latest_timestamp": timestamps[-1] if timestamps else None,
                "time_span_days": None,
                "avg_actions_per_day": None,
            }
        
        # 排序获取最早和最晚时间
        parsed_times.sort()
        earliest = parsed_times[0]
        latest = parsed_times[-1]
        
        # 计算时间跨度（天）
        time_span = (latest - earliest).total_seconds() / (24 * 3600)
        time_span_days = round(time_span, 2)
        
        # 计算平均每天行为数
        if time_span_days > 0:
            avg_actions_per_day = round(len(parsed_times) / time_span_days, 2)
        else:
            # 如果所有行为都在同一天
            avg_actions_per_day = float(len(parsed_times))
        
        return {
            "earliest_timestamp": earliest.strftime("%Y-%m-%d %H:%M:%S"),
            "latest_timestamp": latest.strftime("%Y-%m-%d %H:%M:%S"),
            "time_span_days": time_span_days,
            "avg_actions_per_day": avg_actions_per_day,
        }
    
    def _run_multi_endpoint_evaluation(self, tasks: List[Dict], all_results: List[Dict], 
                                         total_actions: int, update_stats_fn, report_progress_fn,
                                         save_fn=None, error_log_dir=None):
        """
        多端点并行评估（动态负载均衡版本）：
        - 初始按用户哈希分配到端点（利用前缀缓存）
        - 使用工作窃取（Work Stealing）机制动态平衡负载
        - 当某端点任务耗尽时，自动从其他端点获取任务
        - 确保所有端点始终有活可干，最大化资源利用率
        
        Args:
            save_fn: 可选的流式保存函数，用于定期保存中间结果
            error_log_dir: 错误日志目录
        """
        from collections import defaultdict
        
        # 按用户分组任务
        user_tasks = defaultdict(list)
        for task in tasks:
            user_tasks[task["user_id"]].append(task)
        
        # 创建动态任务队列（支持工作窃取）
        task_queue = DynamicTaskQueue(self.endpoints, tasks, dict(user_tasks))
        
        # 共享状态
        lock = threading.Lock()
        results_lock = threading.Lock()
        completed = [0]  # 用列表以便在闭包中修改
        endpoint_first_success = {ep["url"]: False for ep in self.endpoints}  # 每个端点的首次成功标记
        last_progress_milestone = [0]  # 上次输出的进度里程碑 (10%, 20%, ...)
        endpoint_active_workers = {ep["url"]: 0 for ep in self.endpoints}  # 每个端点的活跃工作线程数
        
        def _print_model_output_sample(result: Dict, prefix: str):
            """打印模型输出示例"""
            try:
                from tqdm import tqdm
                tqdm.write(f"\n{'='*60}")
                tqdm.write(f"📝 {prefix}")
                tqdm.write(f"{'='*60}")
                
                # 提取模型输出
                if result.get("model_responses"):
                    # 只显示第一个响应
                    first_response = result["model_responses"][0] if result["model_responses"] else "N/A"
                    tqdm.write(f"   用户ID: {result.get('user_id', 'N/A')}")
                    tqdm.write(f"   行为类型: {result.get('action_type', 'N/A')}")
                    tqdm.write(f"   模型输出: {first_response[:100]}{'...' if len(str(first_response)) > 100 else ''}")
                    if result.get("questions"):
                        q = result["questions"][0]
                        pred_val = q.get('predicted_value', 'N/A')
                        true_val = q.get('true_value', 'N/A')
                        q_type = q.get('type', 'unknown')
                        tqdm.write(f"   问题类型: {q_type}, 预测值: {pred_val}, 真实值: {true_val}")
                else:
                    tqdm.write(f"   结果: {str(result)[:200]}...")
                tqdm.write(f"{'='*60}\n")
            except ImportError:
                print(f"\n[{prefix}] 模型输出: {str(result)[:200]}...\n", flush=True)
        
        def worker_loop(endpoint_url: str, caller: ModelCaller, worker_id: int):
            """
            工作线程循环：持续从队列获取任务直到所有任务完成
            
            这是动态负载均衡的核心：
            1. 优先处理本端点的任务（保持前缀缓存）
            2. 本地任务耗尽后，从其他端点窃取任务
            3. 所有队列都空时退出
            """
            results = []
            
            while True:
                # 从动态队列获取任务（可能是本地任务或窃取的任务）
                task = task_queue.get_task(endpoint_url)
                
                if task is None:
                    # 没有更多任务了，退出循环
                    break
                
                try:
                    # 更新活跃工作线程计数
                    with lock:
                        endpoint_active_workers[endpoint_url] += 1
                    
                    # 执行评估任务
                    result = self.evaluate_single_action(
                        task["user_id"],
                        task["user_profile"],
                        task["action_history"],
                        task["test_action"],
                        model_caller=caller
                    )
                    results.append(result)
                    
                    # 更新统计和进度
                    with lock:
                        completed[0] += 1
                        endpoint_active_workers[endpoint_url] -= 1
                        update_stats_fn(result)
                        task_queue.mark_completed()
                        
                        # 首次成功调用时输出
                        if not endpoint_first_success[endpoint_url] and result.get("success"):
                            endpoint_first_success[endpoint_url] = True
                            _print_model_output_sample(result, f"✅ 端点 [{endpoint_url}] 首次调用成功")
                        
                        # 每 10% 进度输出一次模型响应
                        current_progress = int((completed[0] / total_actions) * 100)
                        current_milestone = (current_progress // 10) * 10
                        if current_milestone > last_progress_milestone[0]:
                            last_progress_milestone[0] = current_milestone
                            _print_model_output_sample(result, f"🚀 进度里程碑: {current_milestone}%")
                        
                        # 定期保存
                        if save_fn:
                            save_fn(results)  # 注意：这里只传增量结果可能不够，save_intermediate_results 需要全量或处理增量
                            # 实际上 save_intermediate_results 接收 list，这里 results 是当前 worker 的积累
                            # 但 all_results 是共享列表，save_intermediate_results 也是基于 len(all_results) 判断
                            # 需要确保 all_results 被正确更新。
                            # 修正：应该在添加到 all_results 后调用 save_fn，并传递 all_results
                
                except Exception as e:
                    print(f"\n任务执行出错: {e}")
                    import traceback
                    traceback.print_exc()
                    with lock:
                        completed[0] += 1
                        endpoint_active_workers[endpoint_url] -= 1
                
                # 定期将结果合并到主列表（减少锁竞争）
                if len(results) >= 10:
                    with results_lock:
                        all_results.extend(results)
                        # 这里调用保存函数更合适
                        if save_fn:
                            save_fn(all_results)
                    results = []
            
            # 循环结束后，处理剩余结果
            if results:
                with results_lock:
                    all_results.extend(results)
                    if save_fn:
                        save_fn(all_results)
            
            return results
        
        def run_endpoint(endpoint_url: str, endpoint_max_workers: int):
            """运行单个端点的工作线程池"""
            # 为此端点创建专用的 ModelCaller
            caller = ModelCaller(self.model_config, base_url_override=endpoint_url, error_log_dir=error_log_dir)
            
            all_endpoint_results = []
            
            # 启动多个工作线程，每个线程独立从队列获取任务
            with ThreadPoolExecutor(max_workers=endpoint_max_workers) as executor:
                futures = []
                for worker_id in range(endpoint_max_workers):
                    future = executor.submit(worker_loop, endpoint_url, caller, worker_id)
                    futures.append(future)
                
                # 等待所有工作线程完成
                for future in as_completed(futures):
                    try:
                        results = future.result()
                        all_endpoint_results.extend(results)
                    except Exception as e:
                        print(f"\n[{endpoint_url}] 工作线程出错: {e}")
            
            return all_endpoint_results
        
        # 所有端点并行执行
        print(f"\n🚀 启动 {len(self.endpoints)} 个端点并行处理（动态负载均衡模式）...")
        print(f"   📌 特性: 工作窃取（Work Stealing）- 空闲端点自动获取其他端点的任务")
        
        with tqdm(total=total_actions, desc="总进度") as pbar:
            # 启动后台线程更新进度条和显示负载均衡状态
            stop_progress = [False]
            
            def update_progress():
                last = 0
                last_stats_print = 0
                while not stop_progress[0] and last < total_actions:
                    with lock:
                        current = completed[0]
                    if current > last:
                        pbar.update(current - last)
                        report_progress_fn(current, total_actions)
                        last = current
                    
                    # 每处理 500 个任务打印一次负载均衡状态
                    if current > 0 and current - last_stats_print >= 500:
                        last_stats_print = current
                        stats = task_queue.get_stats()
                        remaining = stats["remaining_per_endpoint"]
                        stolen = stats["stolen_count"]
                        stolen_rate = stats["stolen_rate"]
                        
                        # 构建状态消息
                        status_parts = []
                        for ep in self.endpoints:
                            url = ep["url"]
                            short_url = url.split("//")[-1].split("/")[0]  # 简化 URL
                            r = remaining.get(url, 0)
                            status_parts.append(f"{short_url}:{r}")
                        
                        try:
                            from tqdm import tqdm
                            tqdm.write(f"\n⚖️  负载均衡状态 [{current}/{total_actions}]: 窃取任务数={stolen} ({stolen_rate:.1f}%)")
                            tqdm.write(f"   剩余任务: {', '.join(status_parts)}")
                        except ImportError:
                            pass
                    
                    time.sleep(0.1)
            
            progress_thread = threading.Thread(target=update_progress, daemon=True)
            progress_thread.start()
            
            # 并行运行所有端点
            with ThreadPoolExecutor(max_workers=len(self.endpoints)) as endpoint_executor:
                endpoint_futures = []
                for ep in self.endpoints:
                    future = endpoint_executor.submit(run_endpoint, ep["url"], ep["max_workers"])
                    endpoint_futures.append(future)
                
                # 等待所有端点完成（结果已经在 worker_loop 中添加到 all_results）
                for future in as_completed(endpoint_futures):
                    try:
                        future.result()  # 结果已在 worker_loop 中收集
                    except Exception as e:
                        print(f"\n端点执行出错: {e}")
            
            # 停止进度更新线程
            stop_progress[0] = True
            progress_thread.join(timeout=1)
            
            # 最终保存
            if save_fn:
                save_fn(all_results)
        
        # 打印负载均衡统计
        stats = task_queue.get_stats()
        print(f"\n📊 动态负载均衡统计:")
        print(f"   总任务数: {stats['total_tasks']}")
        print(f"   完成数: {stats['completed']}")
        print(f"   窃取任务数: {stats['stolen_count']} ({stats['stolen_rate']:.1f}%)")
        print(f"\n   各端点统计:")
        for ep in self.endpoints:
            url = ep["url"]
            ep_stats = stats["endpoint_stats"].get(url, {})
            local = ep_stats.get("local", 0)
            stolen = ep_stats.get("stolen", 0)
            total_processed = local + stolen
            print(f"   [{url}]")
            print(f"      本地任务: {local}, 窃取任务: {stolen}, 总处理: {total_processed}")
            
            # 等待进度条更新完成
            progress_thread.join(timeout=1)
    
    def evaluate_all(self, eval_data: List[Dict], save_intermediate: bool = True, max_workers: int = None) -> List[Dict]:
        """
        评估所有用户的所有待预测行为（滚动预测模式）
        
        历史构建逻辑：
        - 基础历史：来自历史时间范围（如9月份）的所有行为
        - 增量历史：测试时间范围内该测试行为之前的所有行为
        - 完整历史 = 基础历史 + 增量历史
        - 如果超出 token 限制，优先删除时间最早的行为
        
        Args:
            eval_data: load_fixed_experiment_data() 的输出
            save_intermediate: 是否保存中间结果
            max_workers: 最大并发数，None则使用配置文件中的值，1为顺序执行
        
        Returns:
            所有评估结果的列表
        """
        print(f"\n开始评估 {len(eval_data)} 个用户...")
        if len(self.endpoints) > 1:
            total_workers = sum(ep["max_workers"] for ep in self.endpoints)
            print(f"并发数: 多端点模式，总计 {total_workers}")
        else:
            print(f"并发数: {max_workers}")
        print(f"模式: 滚动预测（基础历史 + 测试行为之前的增量历史）")
        if self.history_scene_filter:
            print(f"历史场景过滤: 只使用 [{self.history_scene_filter}] 场景的历史行为")
        if self.max_history_days and self.max_history_days > 0:
            print(f"历史天数过滤: 只检索测试行为前 {self.max_history_days} 天内的历史行为")
        
        all_results = []
        total_actions = sum(len(d["test_actions"]) for d in eval_data)
        
        # 准备所有任务（提前构建每个任务的历史上下文）
        tasks = []
        for user_data in eval_data:
            user_id = user_data["user_id"]
            user_profile = user_data["user_profile"]
            base_history = user_data["base_history"]  # 基础历史（如9月份）
            test_time_all_actions = user_data["test_time_all_actions"]  # 测试时间范围内所有行为
            test_actions = user_data["test_actions"]
            
            # 如果指定了历史场景过滤，对基础历史和增量历史进行过滤
            if self.history_scene_filter:
                filtered_base_history = [
                    action for action in base_history 
                    if action.get("type", "") == self.history_scene_filter
                ]
            else:
                filtered_base_history = base_history
            
            for i, test_item in enumerate(test_actions):
                # 注入 user_id，供历史处理器（RAG/Summary）在需要时做用户级缓存/索引
                test_action = dict(test_item["action"])
                test_action["user_id"] = user_id
                test_time_index = test_item["test_time_index"]  # 在测试时间范围内的索引
                test_action["test_time_index"] = test_time_index # LZQ：用于寻找exp_memory文件
                
                # 构建历史：基础历史 + 测试时间范围内该行为之前的所有行为
                incremental_history = test_time_all_actions[:test_time_index]
                
                # 如果指定了历史场景过滤，对增量历史也进行过滤
                if self.history_scene_filter:
                    filtered_incremental_history = [
                        action for action in incremental_history 
                        if action.get("type", "") == self.history_scene_filter
                    ]
                else:
                    filtered_incremental_history = incremental_history
                
                task_history = filtered_base_history + filtered_incremental_history
                
                # 如果指定了天数限制，对历史行为进行时间过滤
                if self.max_history_days is not None and self.max_history_days > 0:
                    from datetime import datetime, timedelta
                    
                    reference_timestamp = test_action.get("timestamp")
                    if reference_timestamp:
                        try:
                            # 解析参考时间戳
                            if len(reference_timestamp) == 10:  # YYYY-MM-DD
                                ref_datetime = datetime.strptime(reference_timestamp, "%Y-%m-%d")
                            else:  # YYYY-MM-DD HH:MM:SS
                                ref_datetime = datetime.strptime(reference_timestamp[:19], "%Y-%m-%d %H:%M:%S")
                            
                            # 计算截止时间（参考时间点前 N 天）
                            cutoff_datetime = ref_datetime - timedelta(days=self.max_history_days)
                            cutoff_str = cutoff_datetime.strftime("%Y-%m-%d %H:%M:%S")
                            
                            # 过滤掉超出天数范围的行为
                            days_filtered_history = []
                            for action in task_history:
                                action_timestamp = action.get("timestamp", "")
                                if action_timestamp and action_timestamp >= cutoff_str:
                                    days_filtered_history.append(action)
                            
                            task_history = days_filtered_history
                        except ValueError:
                            # 解析失败，跳过天数过滤
                            pass
                
                tasks.append({
                    "user_id": user_id,
                    "user_profile": user_profile,
                    "action_history": task_history,
                    "test_action": test_action,
                    "task_index": i
                })
        
        # 累计 token 统计（用于进度报告）
        cumulative_stats = {
            "prompt_tokens": 0,
            "completion_tokens": 0,
            "cached_tokens": 0,
        }
        last_reported_percentage = 0  # 上次报告的百分比
        last_save_count = [0]  # 上次保存时的结果数量
        
        # 流式保存配置
        SAVE_INTERVAL = 500  # 每完成 500 个任务保存一次
        
        # 延迟创建输出目录（只在第一次保存时创建，避免程序中断留下空文件夹）
        timestamp_output_dir = None
        intermediate_output_path = None
        if save_intermediate:
            date_dir = self.run_date
            timestamp = self.run_timestamp
            timestamp_output_dir = os.path.join(OUTPUT_DIR, date_dir, timestamp)
            intermediate_output_path = os.path.join(timestamp_output_dir, f"{self.experiment_name}_prediction_results.json")
            
            # 设置错误日志目录
            error_log_dir = os.path.join(timestamp_output_dir, "error_logs")
            if not os.path.exists(error_log_dir):
                os.makedirs(error_log_dir, exist_ok=True)
            
            # 如果有单端点 model_caller，设置其错误日志目录
            if self.model_caller:
                self.model_caller.set_error_log_dir(error_log_dir)

            print(f"📁 中间结果将保存到: {intermediate_output_path}")
            print(f"📁 错误日志将保存到: {error_log_dir}")
            print(f"💾 每完成 {SAVE_INTERVAL} 个任务自动保存一次")
        else:
            error_log_dir = None
        
        def update_cumulative_stats(result):
            """更新累计统计"""
            if result.get("binary_mode"):
                cumulative_stats["prompt_tokens"] += result.get("total_prompt_tokens", 0)
                cumulative_stats["completion_tokens"] += result.get("total_completion_tokens", 0)
                cumulative_stats["cached_tokens"] += result.get("total_cached_tokens", 0)
        
        def save_intermediate_results(results_list, force=False):
            """流式保存中间结果（仅在 save_intermediate=True 时生效）"""
            if not save_intermediate or intermediate_output_path is None:
                return
            
            current_count = len(results_list)
            # 每 SAVE_INTERVAL 个结果保存一次，或者强制保存
            if force or (current_count - last_save_count[0] >= SAVE_INTERVAL):
                # 延迟创建目录：只在第一次实际保存时创建
                if not os.path.exists(timestamp_output_dir):
                    os.makedirs(timestamp_output_dir, exist_ok=True)
                    print(f"✅ 创建输出目录: {timestamp_output_dir}")
                try:
                    with open(intermediate_output_path, 'w', encoding='utf-8') as f:
                        json.dump(results_list, f, ensure_ascii=False, indent=2)
                    last_save_count[0] = current_count
                    print(f"\n💾 已保存中间结果: {current_count}/{total_actions} 个任务完成")
                except Exception as e:
                    print(f"\n⚠️ 保存中间结果失败: {e}")
        
        def report_progress(completed_count, total_count):
            """检查并报告进度（每 10% 报告一次）"""
            nonlocal last_reported_percentage
            current_percentage = int((completed_count / total_count) * 10) * 10  # 取整到 10 的倍数
            
            if current_percentage > last_reported_percentage and current_percentage > 0:
                last_reported_percentage = current_percentage
                total_tokens = cumulative_stats["prompt_tokens"] + cumulative_stats["completion_tokens"]
                print(f"\n📊 进度 {current_percentage}% ({completed_count}/{total_count}) Token 统计:")
                print(f"   Prompt Tokens: {cumulative_stats['prompt_tokens']:,}")
                print(f"   Completion Tokens: {cumulative_stats['completion_tokens']:,}")
                print(f"   总 Tokens: {total_tokens:,}")
                # 只有当有缓存数据时才显示缓存相关信息
                if cumulative_stats["cached_tokens"] > 0:
                    print(f"   缓存命中 Tokens: {cumulative_stats['cached_tokens']:,}")
                    if cumulative_stats["prompt_tokens"] > 0:
                        cache_rate = cumulative_stats["cached_tokens"] / cumulative_stats["prompt_tokens"] * 100
                        print(f"   缓存命中率: {cache_rate:.2f}%")
                else:
                    print(f"   缓存命中 Tokens: 无数据（API 未返回缓存信息）")
        
        # 多端点模式：每个端点独立线程池，用户按哈希固定分配
        if len(self.endpoints) > 1:
            self._run_multi_endpoint_evaluation(tasks, all_results, total_actions, 
                                                 update_cumulative_stats, report_progress,
                                                 save_fn=save_intermediate_results,
                                                 error_log_dir=error_log_dir)
        elif max_workers == 1:
            # 单端点顺序执行
            with tqdm(total=total_actions, desc="评估进度") as pbar:
                for task in tasks:
                    result = self.evaluate_single_action(
                        task["user_id"],
                        task["user_profile"],
                        task["action_history"],
                        task["test_action"]
                    )
                    all_results.append(result)
                    update_cumulative_stats(result)
                    pbar.update(1)
                    report_progress(len(all_results), total_actions)
                    save_intermediate_results(all_results)  # 流式保存
        else:
            # 单端点并发执行
            lock = threading.Lock()
            
            def evaluate_task(task):
                return self.evaluate_single_action(
                    task["user_id"],
                    task["user_profile"],
                    task["action_history"],
                    task["test_action"]
                )
            
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                with tqdm(total=total_actions, desc="评估进度") as pbar:
                    future_to_task = {executor.submit(evaluate_task, task): task for task in tasks}
                    
                    for future in as_completed(future_to_task):
                        try:
                            result = future.result()
                            with lock:
                                all_results.append(result)
                                update_cumulative_stats(result)
                                pbar.update(1)
                                report_progress(len(all_results), total_actions)
                                save_intermediate_results(all_results)  # 流式保存
                        except Exception as e:
                            print(f"\n评估任务出错: {e}")
                            with lock:
                                pbar.update(1)
        
        # 统计被过滤的样本
        filtered_count = sum(1 for r in all_results if r.get("filtered", False))
        valid_count = len(all_results) - filtered_count
        
        if filtered_count > 0:
            print(f"\n⚠️  注意: 有 {filtered_count} 个样本因 play_duration 为 0 被过滤")
            print(f"   有效样本数: {valid_count} / {len(all_results)}")
        
        # 保存最终结果和统计信息（按日期和时间分类，精确到秒）
        if save_intermediate:
            # 强制保存最终结果（确保所有结果都被保存）
            save_intermediate_results(all_results, force=True)
            
            # 目录已在流式保存时创建，直接使用
            output_path = intermediate_output_path
            
            # 计算并保存成本统计
            stats = self._calculate_cost_stats(all_results)
            stats_path = os.path.join(timestamp_output_dir, f"{self.experiment_name}_cost_stats.json")
            with open(stats_path, 'w', encoding='utf-8') as f:
                json.dump(stats, f, ensure_ascii=False, indent=2)
            
            # 聚合并保存实际使用的历史行为分布统计
            history_distribution = self._aggregate_history_stats(all_results)
            history_stats_path = os.path.join(timestamp_output_dir, f"{self.experiment_name}_actual_history_distribution.json")
            with open(history_stats_path, 'w', encoding='utf-8') as f:
                json.dump(history_distribution, f, ensure_ascii=False, indent=2)
            
            print(f"\n预测结果已保存到: {output_path}")
            print(f"成本统计已保存到: {stats_path}")
            print(f"实际历史行为分布已保存到: {history_stats_path}")
            print(f"\n💰 Token 使用统计:")
            print(f"  总字符数: {stats['total_chars']:,}")
            print(f"  Prompt字符数: {stats['total_prompt_chars']:,}")
            print(f"  响应字符数: {stats['total_response_chars']:,}")
            if stats.get("token_source") == "api_actual":
                print(f"  📊 真实 Token 统计 (来自 API):")
                print(f"     Prompt Tokens: {stats['total_prompt_tokens']:,}")
                print(f"     Completion Tokens: {stats['total_completion_tokens']:,}")
                print(f"     总 Tokens: {stats['total_tokens']:,}")
                # 只有当有缓存数据时才显示缓存相关信息
                cached_tokens = stats.get('total_cached_tokens', 0)
                if cached_tokens > 0:
                    print(f"     缓存命中 Tokens: {cached_tokens:,}")
                    print(f"     缓存命中率: {stats['cache_hit_rate']:.2%}")
                else:
                    print(f"     缓存命中 Tokens: 无数据（API 未返回缓存信息）")
                print(f"     平均每个问题 Prompt Tokens: {stats['avg_prompt_tokens_per_question']:,}")
                print(f"     平均每个行为 Prompt Tokens: {stats['avg_prompt_tokens_per_action']:,}")
            else:
                print(f"  估算Token数: ~{stats.get('estimated_tokens', 0):,} (API 未返回 usage 信息)")
                cached_tokens = stats.get('total_cached_tokens', 0)
                if cached_tokens > 0:
                    print(f"  缓存命中 Tokens: {cached_tokens:,}")
                else:
                    print(f"  缓存命中 Tokens: 无数据")
            
            # 显示每个 prompt 的 token 长度统计
            prompt_stats = stats.get("prompt_tokens_stats", {})
            if prompt_stats.get("count", 0) > 0:
                print(f"  📏 单个 Prompt Token 长度统计:")
                print(f"     样本数: {prompt_stats['count']:,}")
                print(f"     平均值: {prompt_stats['avg']:,.2f}")
                print(f"     最小值: {prompt_stats['min']:,}")
                print(f"     最大值: {prompt_stats['max']:,}")
                print(f"     中位数: {prompt_stats['median']:,}")
            
            # 生成分布统计和PR曲线数据
            print("\n" + "=" * 80)
            print("生成Behavior分布统计和PR曲线数据...")
            print("=" * 80)
            
            # 预测分布统计
            pred_dist_path = os.path.join(timestamp_output_dir, f"{self.experiment_name}_prediction_distribution.json")
            analyze_prediction_distribution(all_results, pred_dist_path)
            
            # PR曲线数据
            pr_curve_path = os.path.join(timestamp_output_dir, f"{self.experiment_name}_pr_curve_data.json")
            generate_pr_curve_data(all_results, pr_curve_path)
            
            # 生成预测分布可视化图表
            prediction_charts_dir = os.path.join(timestamp_output_dir, "prediction_distribution")
            try:
                from plot.plot_prediction_distribution import generate_prediction_charts
                generate_prediction_charts(all_results, prediction_charts_dir)
            except ImportError:
                print("  ⚠️  plot_prediction_distribution.py not found, skipping prediction chart generation")
            except Exception as e:
                print(f"  ⚠️  Prediction chart generation failed: {e}")
            
            # 生成实际历史行为分布可视化图表（独立目录）
            history_charts_dir = os.path.join(timestamp_output_dir, "actual_history_distribution")
            try:
                from plot.plot_prediction_distribution import generate_actual_history_charts
                generate_actual_history_charts(all_results, history_charts_dir)
            except ImportError:
                print("  ⚠️  plot_prediction_distribution.py not found, skipping history chart generation")
            except Exception as e:
                print(f"  ⚠️  History chart generation failed: {e}")
            
            # 生成结果分析图表（分布对比 + AUC/F1曲线）
            results_analysis_dir = os.path.join(timestamp_output_dir, "results_analysis")
            try:
                from plot.plot_results_analysis import generate_results_analysis
                generate_results_analysis(all_results, results_analysis_dir)
            except ImportError:
                print("  ⚠️  plot_results_analysis.py not found, skipping results analysis chart generation")
            except Exception as e:
                print(f"  ⚠️  Results analysis chart generation failed: {e}")
            
            # 如果有测试数据分布统计，生成对比
            # 尝试加载测试数据分布文件（从实验数据路径推断）
            try:
                if USE_FIXED_EXPERIMENT_DATA and os.path.exists(FIXED_EXPERIMENT_DATA_PATH):
                    test_dist_path = FIXED_EXPERIMENT_DATA_PATH.replace(".json", "_test_distribution.json")
                    if os.path.exists(test_dist_path):
                        print(f"\n找到测试数据分布文件，生成对比...")
                        with open(test_dist_path, 'r', encoding='utf-8') as f:
                            test_distribution = json.load(f)
                        with open(pred_dist_path, 'r', encoding='utf-8') as f:
                            pred_distribution = json.load(f)
                        
                        comparison_path = os.path.join(timestamp_output_dir, f"{self.experiment_name}_distribution_comparison.json")
                        generate_distribution_comparison(test_distribution, pred_distribution, comparison_path)
            except Exception as e:
                print(f"\n生成分布对比时出错（非致命）: {e}")
        
        return all_results
    
    def _calculate_cost_stats(self, results: List[Dict]) -> Dict:
        """计算成本统计信息（支持新旧两种数据格式，包含真实 token 统计）"""
        total_prompt_chars = 0
        total_response_chars = 0
        total_prompt_tokens = 0  # API 返回的真实 prompt token 数
        total_completion_tokens = 0  # API 返回的真实 completion token 数
        total_cached_tokens = 0  # 缓存命中的 token 数
        successful_calls = 0
        total_questions = 0
        failed_questions = 0
        total_retries = 0
        
        # 收集每个 question 的 prompt tokens，用于计算统计
        all_prompt_tokens = []
        
        for result in results:
            if result.get("success"):
                successful_calls += 1
                
                # 新格式（二分类模式）
                if result.get("binary_mode"):
                    total_prompt_chars += result.get("total_prompt_length", 0)
                    total_response_chars += result.get("total_response_length", 0)
                    total_prompt_tokens += result.get("total_prompt_tokens", 0)
                    total_completion_tokens += result.get("total_completion_tokens", 0)
                    total_cached_tokens += result.get("total_cached_tokens", 0)
                    total_questions += len(result.get("questions", []))
                    failed_questions += len(result.get("failed_questions", []))
                    
                    # 统计重试次数和收集每个 prompt 的 tokens
                    for q in result.get("questions", []):
                        total_retries += q.get("retry_count", 0)
                        # 收集每个 question 的 prompt tokens
                        prompt_tokens = q.get("prompt_tokens", 0)
                        if prompt_tokens > 0:
                            all_prompt_tokens.append(prompt_tokens)
                    for q in result.get("failed_questions", []):
                        total_retries += q.get("retry_count", 0)
                        prompt_tokens = q.get("prompt_tokens", 0)
                        if prompt_tokens > 0:
                            all_prompt_tokens.append(prompt_tokens)
                else:
                    # 旧格式（传统模式）
                    total_prompt_chars += result.get("prompt_length", 0)
                    total_response_chars += result.get("response_length", 0)
                    total_questions += len(result.get("questions", []))
        
        total_chars = total_prompt_chars + total_response_chars
        total_tokens = total_prompt_tokens + total_completion_tokens
        
        # 如果有真实 token 数据，使用真实数据；否则使用估算
        # 粗略估算 token 数（中文约 1.5 字符/token，英文约 4 字符/token）
        # 这里使用平均值 2.5 字符/token
        estimated_tokens = int(total_chars / 2.5) if total_tokens == 0 else None
        
        stats = {
            "model_name": self.model_name,
            "successful_calls": successful_calls,
            "total_prompt_chars": total_prompt_chars,
            "total_response_chars": total_response_chars,
            "total_chars": total_chars,
            "avg_prompt_chars": int(total_prompt_chars / successful_calls) if successful_calls > 0 else 0,
            "avg_response_chars": int(total_response_chars / successful_calls) if successful_calls > 0 else 0,
            "total_questions_predicted": total_questions,
            "failed_questions": failed_questions,
            "total_retries": total_retries,
        }
        
        # 添加 token 统计（优先使用 API 返回的真实数据）
        if total_tokens > 0:
            stats["total_prompt_tokens"] = total_prompt_tokens
            stats["total_completion_tokens"] = total_completion_tokens
            stats["total_tokens"] = total_tokens
            stats["total_cached_tokens"] = total_cached_tokens
            # 注意：avg 使用问题数（total_questions）作为分母，因为每个问题有一个 prompt
            stats["avg_prompt_tokens_per_question"] = int(total_prompt_tokens / total_questions) if total_questions > 0 else 0
            stats["avg_completion_tokens_per_question"] = int(total_completion_tokens / total_questions) if total_questions > 0 else 0
            # 保留按行为数计算的平均值（用于成本估算）
            stats["avg_prompt_tokens_per_action"] = int(total_prompt_tokens / successful_calls) if successful_calls > 0 else 0
            stats["cache_hit_rate"] = total_cached_tokens / total_prompt_tokens if total_prompt_tokens > 0 else 0
            stats["token_source"] = "api_actual"  # 标记数据来源为 API 真实值
        else:
            stats["estimated_tokens"] = estimated_tokens
            stats["total_cached_tokens"] = total_cached_tokens
            stats["token_source"] = "estimated"  # 标记数据来源为估算值
        
        # 添加每个 prompt 的 token 长度统计（平均、最小、最大）
        if all_prompt_tokens:
            stats["prompt_tokens_stats"] = {
                "count": len(all_prompt_tokens),
                "avg": round(sum(all_prompt_tokens) / len(all_prompt_tokens), 2),
                "min": min(all_prompt_tokens),
                "max": max(all_prompt_tokens),
                "median": sorted(all_prompt_tokens)[len(all_prompt_tokens) // 2],
            }
        else:
            stats["prompt_tokens_stats"] = {
                "count": 0,
                "avg": 0,
                "min": 0,
                "max": 0,
                "median": 0,
            }
        
        return stats
    
    def _aggregate_history_stats(self, results: List[Dict]) -> Dict:
        """
        聚合所有预测结果的实际历史行为统计
        
        Args:
            results: 所有预测结果列表
            
        Returns:
            聚合后的历史行为分布统计
        """
        # 聚合场景分布
        total_scene_counts = {}
        # 聚合行为分布
        total_action_counts = {}
        # 统计信息
        total_samples = 0
        total_original_count = 0
        total_filtered_count = 0
        total_actual_used_count = 0
        total_actual_used_tokens = 0
        
        # prompt tokens 统计
        total_prompt_tokens = 0
        total_prompt_chars = 0
        
        # 用户级别统计
        user_stats = {}  # user_id -> {"history_count": [], "history_tokens": [], "prompt_tokens": []}
        
        for result in results:
            if result.get("filtered", False):
                continue
            
            history_stats = result.get("history_stats", {})
            if not history_stats:
                continue
            
            total_samples += 1
            total_original_count += history_stats.get("original_count", 0)
            total_filtered_count += history_stats.get("filtered_count", 0)
            total_actual_used_count += history_stats.get("actual_used_count", 0)
            total_actual_used_tokens += history_stats.get("actual_used_tokens", 0)
            
            # 收集 prompt tokens 统计
            prompt_tokens = result.get("total_prompt_tokens", 0)
            prompt_chars = result.get("total_prompt_length", 0)
            total_prompt_tokens += prompt_tokens
            total_prompt_chars += prompt_chars
            
            # 按用户聚合统计
            user_id = result.get("user_id", "unknown")
            if user_id not in user_stats:
                user_stats[user_id] = {
                    "history_count": [],
                    "history_tokens": [],
                    "prompt_tokens": [],
                    "prompt_chars": []
                }
            user_stats[user_id]["history_count"].append(history_stats.get("actual_used_count", 0))
            user_stats[user_id]["history_tokens"].append(history_stats.get("actual_used_tokens", 0))
            user_stats[user_id]["prompt_tokens"].append(prompt_tokens)
            user_stats[user_id]["prompt_chars"].append(prompt_chars)
            
            # 聚合场景分布
            for scene, count in history_stats.get("scene_distribution", {}).items():
                total_scene_counts[scene] = total_scene_counts.get(scene, 0) + count
            
            # 聚合行为分布
            for action, count in history_stats.get("action_distribution", {}).items():
                total_action_counts[action] = total_action_counts.get(action, 0) + count
        
        # 计算场景分布百分比
        total_scene_actions = sum(total_scene_counts.values())
        scene_distribution = {}
        for scene, count in sorted(total_scene_counts.items(), key=lambda x: -x[1]):
            scene_distribution[scene] = {
                "count": count,
                "percentage": round(count / total_scene_actions * 100, 2) if total_scene_actions > 0 else 0
            }
        
        # 计算行为分布百分比
        total_actions = sum(total_action_counts.values())
        action_distribution = {}
        for action, count in sorted(total_action_counts.items(), key=lambda x: -x[1]):
            action_distribution[action] = {
                "count": count,
                "percentage": round(count / total_actions * 100, 2) if total_actions > 0 else 0
            }
        
        # 按场景分组的行为分布
        per_scene_actions = {}
        for action_key, count in total_action_counts.items():
            if "_" in action_key:
                scene = action_key.rsplit("_", 1)[0]
                # 处理可能的中文场景名带下划线的情况
                for known_scene in total_scene_counts.keys():
                    if action_key.startswith(known_scene + "_"):
                        scene = known_scene
                        break
                if scene not in per_scene_actions:
                    per_scene_actions[scene] = {}
                per_scene_actions[scene][action_key] = count
        
        # 计算用户级别的统计
        num_users = len(user_stats)
        user_avg_history_count = []
        user_avg_history_tokens = []
        user_avg_prompt_tokens = []
        user_avg_prompt_chars = []
        
        for user_id, stats in user_stats.items():
            if stats["history_count"]:
                user_avg_history_count.append(sum(stats["history_count"]) / len(stats["history_count"]))
                user_avg_history_tokens.append(sum(stats["history_tokens"]) / len(stats["history_tokens"]))
                user_avg_prompt_tokens.append(sum(stats["prompt_tokens"]) / len(stats["prompt_tokens"]))
                user_avg_prompt_chars.append(sum(stats["prompt_chars"]) / len(stats["prompt_chars"]))
        
        return {
            "metadata": {
                "total_samples": total_samples,
                "total_users": num_users,
                "total_original_history_count": total_original_count,
                "total_filtered_history_count": total_filtered_count,
                "total_actual_used_history_count": total_actual_used_count,
                "total_actual_used_tokens": total_actual_used_tokens,
                # 样本级别平均值
                "avg_history_per_sample": round(total_actual_used_count / total_samples, 2) if total_samples > 0 else 0,
                "avg_history_tokens_per_sample": round(total_actual_used_tokens / total_samples, 2) if total_samples > 0 else 0,
                # prompt 统计
                "total_prompt_tokens": total_prompt_tokens,
                "avg_prompt_tokens_per_sample": round(total_prompt_tokens / total_samples, 2) if total_samples > 0 else 0,
                # 用户级别平均值（每个用户的平均值的平均）
                "avg_history_per_user": round(sum(user_avg_history_count) / len(user_avg_history_count), 2) if user_avg_history_count else 0,
                "avg_history_tokens_per_user": round(sum(user_avg_history_tokens) / len(user_avg_history_tokens), 2) if user_avg_history_tokens else 0,
                "avg_prompt_tokens_per_user": round(sum(user_avg_prompt_tokens) / len(user_avg_prompt_tokens), 2) if user_avg_prompt_tokens else 0,
            },
            "scene_distribution": scene_distribution,
            "action_distribution": action_distribution,
            "per_scene_actions": per_scene_actions,
        }
    
    def calculate_metrics(self, results: List[Dict], compute_bertscore: bool = True) -> Dict:
        """
        根据评估结果计算所有指标（支持新旧两种数据格式）
        
        Args:
            results: 评估结果列表
            compute_bertscore: 是否计算 BERTScore（需要 GPU，较慢）
        
        Returns:
            {
                "binary_metrics": {...},
                "continuous_metrics": {...},
                "text_metrics": {...},
                "overall": {...}
            }
        """

    def _get_bertscore_flag(self, result: Dict, question: Dict) -> bool:
        """
        确定是否应该计算 BERTScore
        1. 如果 question 中已有 has_prior_user_speech 标志，直接使用 (True/False)
        2. 如果是旧数据（无标志），且是电商客服对话场景，则根据 context 重新计算
        3. 其他场景默认返回 True
        """
        # 1. 优先使用 existing flag
        if "has_prior_user_speech" in question:
            return question["has_prior_user_speech"]
        
        # 2. 如果是旧数据，检查场景
        action_type = result.get("action_type")
        # 兼容旧数据的 action_type 获取方式
        if not action_type and "scene_info" in result:
             action_type = result["scene_info"].get("type")
             
        if action_type != "电商客服对话":
            return True
            
        # 3. 电商客服对话：检查历史中是否有用户发言
        true_value = question.get("true_value", "")
        # 注意：true_value 可能是非字符串类型，转为 string
        if true_value is not None:
            true_value = str(true_value)
            
        scene_info = result.get("scene_info", {})
        actions = scene_info.get("action", [])
        
        dialogue_content = []
        for act in actions:
            if act.get("type") == "dialogue":
                dialogue_content = act.get("content", [])
                break
        
        if not dialogue_content:
            return True # 无法判断，默认计算
            
        # 倒序查找 target message（因为 prompt_builder 也是取最后一个符合条件的）
        target_idx = -1
        for i in range(len(dialogue_content) - 1, -1, -1):
            msg = dialogue_content[i]
            if msg.get("role") == "user":
                content = msg.get("content", "")
                # 尝试精确匹配
                if content == true_value:
                    target_idx = i
                    break
                    
        if target_idx == -1:
            # 如果没找到完全匹配的，尝试找包含关系或默认 True
            # print(f"Warning: Could not find target message '{true_value}' in dialogue. Defaulting to True.")
            return True
            
        # 检查 target 之前是否有 user 消息
        has_prior_user = False
        for i in range(target_idx):
            msg = dialogue_content[i]
            if msg.get("role") == "user":
                has_prior_user = True
                break
                
        return has_prior_user

    def calculate_metrics(self, results: List[Dict], compute_bertscore: bool = True) -> Dict:
        """
        计算所有评估指标
        
        Args:
            results: 预测结果列表
            compute_bertscore: 是否计算 BERTScore
            
        Returns:
            Dict: 包含各类指标的字典
        """
        print("\n计算评估指标...")
        
        # 按字段收集数据
        binary_data = {}  # field -> {"y_true": [], "y_pred_labels": [], "y_pred_probs": []}
        continuous_data = {}  # field -> (y_true, y_pred)
        text_data = {}  # field -> (references, hypotheses, (new) bertscore_flags)
        
        successful_api_calls = 0
        total_questions = 0
        failed_parsing_count = 0
        filtered_samples = 0
        failed_binary_questions = 0  # 二分类模式下失败的问题数
        total_retries = 0  # 总重试次数
        
        # 用于统计预测方法
        logprobs_predictions = 0
        direct_mapping_predictions = 0
        
        for result in results:
            if result.get("filtered", False):
                filtered_samples += 1
                continue
            
            if result["success"]:
                successful_api_calls += 1
                
                # 统计失败的问题（二分类模式）
                if result.get("binary_mode"):
                    failed_binary_questions += len(result.get("failed_questions", []))
                    for q in result.get("failed_questions", []):
                        total_retries += q.get("retry_count", 0)
                
                for q in result["questions"]:
                    total_questions += 1
                    field = q["field"]
                    q_type = q["type"]
                    true_val = q["true_value"]
                    pred_val = q["predicted_value"]
                    
                    # 统计预测方法
                    if q.get("prediction_method") == "logprobs":
                        logprobs_predictions += 1
                    elif q.get("prediction_method") == "direct_mapping":
                        direct_mapping_predictions += 1
                    
                    # 统计重试次数
                    total_retries += q.get("retry_count", 0)
                    
                    # 如果预测值为None，说明解析失败，跳过该样本
                    if pred_val is None:
                        failed_parsing_count += 1
                        continue
                    
                    if q_type == "binary":
                        # 排除不参与评估的字段
                        if field in {"video_downloaded", "ad_converted"}:
                            continue
                        if field not in binary_data:
                            binary_data[field] = {"y_true": [], "y_pred_labels": [], "y_pred_probs": []}
                        try:
                            # 获取预测标签（基于YES/NO直接判定）和概率
                            predicted_label = q.get("predicted_label")
                            if predicted_label is None:
                                # 兼容旧数据：如果没有predicted_label，从prediction推断
                                predicted_label = 1 if float(pred_val) >= 0.5 else 0
                            
                            binary_data[field]["y_true"].append(int(true_val))
                            binary_data[field]["y_pred_labels"].append(int(predicted_label))
                            binary_data[field]["y_pred_probs"].append(float(pred_val))
                        except (ValueError, TypeError):
                            failed_parsing_count += 1
                    
                    elif q_type == "continuous":
                        if field not in continuous_data:
                            continuous_data[field] = {"y_true": [], "y_pred": [], "normalizers": []}
                        try:
                            continuous_data[field]["y_true"].append(float(true_val))
                            continuous_data[field]["y_pred"].append(float(pred_val))
                            # 收集归一化因子（如视频时长），用于计算 NMAE
                            normalizer = q.get("video_duration")
                            # 兼容旧数据：尝试从 result 的 scene_info.context 中提取 duration
                            if normalizer is None and field == "video_watch_seconds":
                                scene_info = result.get("scene_info", {})
                                context = scene_info.get("context", {})
                                duration_raw = context.get("duration")
                                if duration_raw is not None:
                                    try:
                                        if isinstance(duration_raw, (int, float)):
                                            normalizer = float(duration_raw) if duration_raw > 0 else None
                                        elif isinstance(duration_raw, str):
                                            val = float(duration_raw.replace("秒", "").strip()) if duration_raw else 0
                                            normalizer = val if val > 0 else None
                                    except (ValueError, TypeError):
                                        normalizer = None
                            if normalizer is not None:
                                try:
                                    continuous_data[field]["normalizers"].append(float(normalizer))
                                except (ValueError, TypeError):
                                    continuous_data[field]["normalizers"].append(None)
                            else:
                                continuous_data[field]["normalizers"].append(None)
                        except (ValueError, TypeError):
                            failed_parsing_count += 1
                    
                    elif q_type == "text":
                        if field not in text_data:
                            text_data[field] = {"references": [], "hypotheses": [], "bertscore_flags": [], "questions": []}
                        text_data[field]["references"].append(str(true_val))
                        text_data[field]["hypotheses"].append(str(pred_val))
                        text_data[field]["questions"].append(q)
                        
                        # 收集 BERTScore 计算标志
                        should_calc = self._get_bertscore_flag(result, q)
                        text_data[field]["bertscore_flags"].append(should_calc)
        
        # 计算二分类指标
        binary_metrics = {}
        for field, data in binary_data.items():
            print(f"  计算二分类指标: {field} (样本数: {len(data['y_true'])})")
            binary_metrics[field] = calculate_all_binary_metrics(
                data["y_true"], 
                data["y_pred_labels"],  # 基于YES/NO直接判定的标签
                data["y_pred_probs"],   # 概率值，用于LogLoss/AUC/ECE
                field_name=field
            )
        
        # 计算连续值指标
        continuous_metrics = {}
        for field, data in continuous_data.items():
            print(f"  计算连续值指标: {field} (样本数: {len(data['y_true'])})")
            # 检查是否有有效的归一化因子
            normalizers = data.get("normalizers", [])
            has_valid_normalizers = any(n is not None and n > 0 for n in normalizers)
            continuous_metrics[field] = calculate_all_continuous_metrics(
                data["y_true"], data["y_pred"],
                normalizers=normalizers if has_valid_normalizers else None
            )
        
        # 计算文本指标
        text_metrics = {}
        for field, data in text_data.items():
            bertscore_info = " (含BERTScore)" if compute_bertscore else ""
            print(f"  计算文本指标: {field} (样本数: {len(data['references'])}){bertscore_info}")
            
            metrics_result = calculate_all_text_metrics(
                data["references"], data["hypotheses"],
                compute_bertscore=compute_bertscore,
                bertscore_flags=data.get("bertscore_flags")
            )
            text_metrics[field] = metrics_result
            
            # 将 Per-sample BERTScore 注入到 question 对象中
            if compute_bertscore:
                f1_list = metrics_result.get("BERTScore_F1_list", [])
                p_list = metrics_result.get("BERTScore_P_list", [])
                r_list = metrics_result.get("BERTScore_R_list", [])
                
                questions = data.get("questions", [])
                
                # 确保列表长度一致
                if len(f1_list) == len(questions):
                    for i, q in enumerate(questions):
                        # 如果没有计算 BERTScore（None 或 NaN），则不添加字段或设为 None
                        f1_val = f1_list[i]
                        if f1_val is not None and not (isinstance(f1_val, float) and np.isnan(f1_val)):
                            q["bert_score"] = {
                                "P": p_list[i],
                                "R": r_list[i],
                                "F1": f1_list[i]
                            }
                        else:
                            # 即使没有计算，也可以明确标记为 None，或者直接不加字段
                            # 这里选择不加字段，或者根据需求添加
                            pass
                else:
                    print(f"Warning: BERTScore list length ({len(f1_list)}) matches question length ({len(questions)}) mismatch for field {field}")
        
        # 计算 Micro 和 Macro F1
        micro_macro_f1 = calculate_micro_macro_f1(binary_metrics)
        print(f"  Micro F1: {micro_macro_f1['Micro_F1']:.4f}, Macro F1: {micro_macro_f1['Macro_F1']:.4f}")
        
        # 整体统计
        valid_samples = len(results) - filtered_samples
        overall = {
            "total_predictions": len(results),
            "filtered_samples": filtered_samples,
            "valid_samples": valid_samples,
            "successful_api_calls": successful_api_calls,
            "api_success_rate": successful_api_calls / valid_samples if valid_samples > 0 else 0,
            "total_questions": total_questions,
            "failed_parsing_questions": failed_parsing_count,
            "failed_binary_questions": failed_binary_questions,
            "total_retries": total_retries,
            "parsing_success_rate": 1 - (failed_parsing_count / total_questions) if total_questions else 0,
            "logprobs_predictions": logprobs_predictions,
            "direct_mapping_predictions": direct_mapping_predictions,
            # 添加 Micro/Macro 汇总指标
            "Micro_F1": micro_macro_f1["Micro_F1"],
            "Macro_F1": micro_macro_f1["Macro_F1"],
            "Micro_Precision": micro_macro_f1["Micro_Precision"],
            "Micro_Recall": micro_macro_f1["Micro_Recall"],
            "Total_TP": micro_macro_f1["Total_TP"],
            "Total_FP": micro_macro_f1["Total_FP"],
            "Total_FN": micro_macro_f1["Total_FN"],
        }
        
        # 统计 LLM Judge Metrics
        llm_judge_data = {
            "intent_fidelity": [],
            "persona_mimicry": [],
            "knowledge_boundary": [],
            "semantic_alignment": [],
            "average_score": []
        }
        
        has_llm_judge_scores = False
        
        for result in results:
            if "llm_judge_scores" in result:
                has_llm_judge_scores = True
                scores = result["llm_judge_scores"]
                llm_judge_data["intent_fidelity"].append(scores.get("intent_fidelity", 0))
                llm_judge_data["persona_mimicry"].append(scores.get("persona_mimicry", 0))
                llm_judge_data["knowledge_boundary"].append(scores.get("knowledge_boundary", 0))
                llm_judge_data["semantic_alignment"].append(scores.get("semantic_alignment", 0))
                llm_judge_data["average_score"].append(result.get("llm_judge_avg_score", 0))

        llm_judge_metrics = {}
        if has_llm_judge_scores:
            print(f"  计算 LLM Judge 指标 (样本数: {len(llm_judge_data['average_score'])})")
            for key, values in llm_judge_data.items():
                if values:
                    llm_judge_metrics[key] = sum(values) / len(values)
                else:
                    llm_judge_metrics[key] = 0.0

        return {
            "binary_metrics": binary_metrics,
            "continuous_metrics": continuous_metrics,
            "text_metrics": text_metrics,
            "llm_judge_metrics": llm_judge_metrics,
            "overall": overall,
        }
    
    def save_metrics_report(self, metrics: Dict, output_path: str):
        """保存指标报告（按日期和时间分类，精确到秒，数值保留两位小数）"""
        # 使用统一的运行日期和时间戳
        date_dir = self.run_date
        timestamp = self.run_timestamp
        
        # 修改输出路径，添加日期/时间两级子目录
        output_dir = os.path.dirname(output_path)
        filename = os.path.basename(output_path)
        timestamp_output_dir = os.path.join(output_dir, date_dir, timestamp)
        output_path = os.path.join(timestamp_output_dir, filename)
        
        os.makedirs(timestamp_output_dir, exist_ok=True)
        
        # 先将百分比指标转换为百分数形式（0.23 -> 23.0）
        metrics_percentage = convert_metrics_to_percentage(metrics)
        
        # 将浮点数保留两位小数
        metrics_rounded = round_floats(metrics_percentage, decimals=2)
        
        # JSON格式
        json_path = output_path.replace(".txt", ".json")
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(metrics_rounded, f, ensure_ascii=False, indent=2)
        
        # 可读文本格式
        lines = []
        lines.append("=" * 80)
        lines.append("用户模拟评估报告")
        lines.append("=" * 80)
        lines.append("")
        
        # 整体统计
        lines.append("## 整体统计")
        lines.append("-" * 80)
        
        # 按重要性排序显示统计信息
        key_order = [
            "total_predictions", "filtered_samples", "valid_samples",
            "successful_api_calls", "api_success_rate",
            "total_questions", "failed_parsing_questions", "failed_binary_questions",
            "total_retries", "parsing_success_rate",
            "logprobs_predictions", "direct_mapping_predictions",
            # Micro/Macro 汇总指标
            "Micro_F1", "Macro_F1", "Micro_Precision", "Micro_Recall",
            "Total_TP", "Total_FP", "Total_FN"
        ]
        
        # 先按顺序显示已定义的键
        for key in key_order:
            if key in metrics_rounded["overall"]:
                value = metrics_rounded["overall"][key]
                if isinstance(value, float):
                    lines.append(f"{key:35s}: {value:.2f}")
                elif value is None:
                    lines.append(f"{key:35s}: N/A")
                else:
                    lines.append(f"{key:35s}: {value}")
        
        # 再显示其他键
        for key, value in metrics_rounded["overall"].items():
            if key not in key_order:
                if isinstance(value, float):
                    lines.append(f"{key:35s}: {value:.2f}")
                elif value is None:
                    lines.append(f"{key:35s}: N/A")
                else:
                    lines.append(f"{key:35s}: {value}")
        lines.append("")
        
        # 二分类指标
        if metrics_rounded["binary_metrics"]:
            lines.append("## 二分类预测指标")
            lines.append("-" * 80)
            
            for field, field_metrics in metrics_rounded["binary_metrics"].items():
                lines.append(f"\n### {field}")
                for metric, value in field_metrics.items():
                    if isinstance(value, float):
                        lines.append(f"  {metric:20s}: {value:.2f}")
                    elif value is None:
                        lines.append(f"  {metric:20s}: N/A")
                    else:
                        lines.append(f"  {metric:20s}: {value}")
            lines.append("")
        
        # 连续值指标
        if metrics_rounded["continuous_metrics"]:
            lines.append("## 连续值预测指标")
            lines.append("-" * 80)
            for field, field_metrics in metrics_rounded["continuous_metrics"].items():
                lines.append(f"\n### {field}")
                for metric, value in field_metrics.items():
                    if isinstance(value, float):
                        lines.append(f"  {metric:20s}: {value:.2f}")
                    elif value is None:
                        lines.append(f"  {metric:20s}: N/A")
                    else:
                        lines.append(f"  {metric:20s}: {value}")
            lines.append("")
        
        # 文本指标
        if metrics_rounded["text_metrics"]:
            lines.append("## 文本预测指标")
            lines.append("-" * 80)
            for field, field_metrics in metrics_rounded["text_metrics"].items():
                lines.append(f"\n### {field}")
                for metric, value in field_metrics.items():
                    if isinstance(value, float):
                        lines.append(f"  {metric:20s}: {value:.2f}")
                    elif value is None:
                        lines.append(f"  {metric:20s}: N/A")
                    else:
                        lines.append(f"  {metric:20s}: {value}")
            lines.append("")
        
        # LLM Judge 指标
        if metrics_rounded.get("llm_judge_metrics"):
            lines.append("## LLM Judge 评估指标 (电商客服对话)")
            lines.append("-" * 80)
            
            judge_metrics = metrics_rounded["llm_judge_metrics"]
            # 顺序: Average, Intent, Persona, Knowledge, Semantic
            key_order = ["average_score", "intent_fidelity", "persona_mimicry", "knowledge_boundary", "semantic_alignment"]
            
            for key in key_order:
                if key in judge_metrics:
                    value = judge_metrics[key]
                    display_key = key.replace("_", " ").title()
                    if isinstance(value, float):
                        lines.append(f"  {display_key:25s}: {value:.2f}")
                    elif value is None:
                        lines.append(f"  {display_key:25s}: N/A")
                    else:
                        lines.append(f"  {display_key:25s}: {value}")
            lines.append("")
        
        lines.append("=" * 80)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("\n".join(lines))
        
        # 生成 Excel 文件
        excel_path = output_path.replace(".txt", ".xlsx")
        try:
            generate_metrics_excel(metrics_rounded, excel_path)
        except Exception as e:
            print(f"  - Excel 生成失败: {e}")
        
        print(f"\n指标报告已保存到:")
        print(f"  - JSON: {json_path}")
        print(f"  - TXT:  {output_path}")


# 需要跳过的场景类型（评估时跳过这些场景的单个行为，其他行为正常评估）
# 注：电商客服对话已支持，不再跳过
SKIP_SCENE_TYPES = set()


def load_fixed_experiment_data(data_path: str) -> Tuple[List[Dict], Dict]:
    """
    从固定的实验数据文件加载数据
    
    新数据结构支持分离时间范围的滚动预测：
    - base_history: 基础历史行为（如9月份）
    - test_time_all_actions: 测试时间范围内的所有行为（用于滚动预测增量历史）
    - test_actions: 被采样的待预测行为（包含 test_time_index）
    
    注意：会自动过滤掉 test_actions 中属于 SKIP_SCENE_TYPES 的单个行为
    
    Returns:
        Tuple of (eval_data, metadata)
        - eval_data: List of evaluation data
        - metadata: 实验数据的元数据
    """
    print(f"从固定数据集加载: {data_path}")
    
    with open(data_path, 'r', encoding='utf-8') as f:
        experiment_data = json.load(f)
    
    metadata = experiment_data.get("metadata", {})
    users = experiment_data.get("users", [])
    
    print(f"  元数据:")
    print(f"    用户数: {metadata.get('num_users')}")
    print(f"    测试集长度: {metadata.get('test_length')}")
    print(f"    总测试行为: {metadata.get('total_test_actions')}")
    print(f"    创建时间: {metadata.get('created_at')}")
    
    max_history_tokens = metadata.get('max_history_tokens')
    if max_history_tokens:
        print(f"    历史行为最大Token数: {max_history_tokens}")
    
    # 显示历史行为token统计（如果有）
    if 'history_token_statistics' in metadata:
        token_stats = metadata['history_token_statistics']
        print(f"    基础历史Token统计:")
        print(f"      平均每用户: {token_stats.get('avg_tokens_per_user', 0):,.0f}")
        print(f"      最小: {token_stats.get('min_tokens', 0):,}")
        print(f"      最大: {token_stats.get('max_tokens', 0):,}")
    
    # 显示时间分布（如果有）
    if 'time_distribution' in metadata:
        time_dist = metadata['time_distribution']
        print(f"    时间分布区间:")
        if 'history' in time_dist:
            print(f"      基础历史: {time_dist['history'].get('earliest', 'N/A')} ~ {time_dist['history'].get('latest', 'N/A')}")
        if 'test' in time_dist:
            print(f"      测试集: {time_dist['test'].get('earliest', 'N/A')} ~ {time_dist['test'].get('latest', 'N/A')}")
    
    # 过滤掉 test_actions 中属于需要跳过场景的单个行为（保留其他行为）
    total_skipped_actions = 0
    skipped_actions_by_type = {}
    
    # 转换为evaluator需要的格式，同时过滤掉需要跳过的行为
    eval_data = []
    total_original_actions = 0
    total_valid_actions = 0
    
    for user in users:
        # 过滤掉需要跳过的场景类型的行为
        filtered_test_actions = []
        for t in user["test_actions"]:
            action = t["action"]
            action_type = action.get("type", "")
            total_original_actions += 1
            
            if action_type in SKIP_SCENE_TYPES:
                # 跳过该行为，记录统计
                total_skipped_actions += 1
                skipped_actions_by_type[action_type] = skipped_actions_by_type.get(action_type, 0) + 1
            else:
                # 保留该行为
                filtered_test_actions.append({
                    "action": action,
                    "test_time_index": t["test_time_index"]
                })
                total_valid_actions += 1
        
        # 只有当用户有有效的测试行为时才添加
        if filtered_test_actions:
            eval_data.append({
                "user_id": user["user_id"],
                "user_profile": user["user_profile"],
                "base_history": user["base_history"],
                "test_time_all_actions": user["test_time_all_actions"],
                "test_actions": filtered_test_actions
            })
    
    # 打印跳过的行为信息
    if total_skipped_actions > 0:
        print(f"\n⚠️  已过滤 {total_skipped_actions} 个特殊场景行为:")
        print(f"    跳过的场景类型: {SKIP_SCENE_TYPES}")
        for scene_type, count in skipped_actions_by_type.items():
            print(f"      {scene_type}: {count} 个行为")
        print(f"    有效行为数: {total_valid_actions} / {total_original_actions}")
    
    print(f"✓ 成功加载 {len(eval_data)} 个用户的实验数据（共 {total_valid_actions} 个待评估行为）")
    
    # 更新元数据，记录跳过的行为
    metadata["skipped_actions_count"] = total_skipped_actions
    metadata["skipped_actions_by_type"] = skipped_actions_by_type
    metadata["skipped_scene_types"] = list(SKIP_SCENE_TYPES)
    metadata["actual_users_count"] = len(eval_data)
    metadata["actual_test_actions_count"] = total_valid_actions
    
    return eval_data, metadata


def main():
    """主评估流程 - 支持多模型评估"""
    import argparse
    
    parser = argparse.ArgumentParser(description="用户模拟评估系统 - 多模型支持（滚动预测）")
    parser.add_argument("--use-fixed-data", type=str, default="./work_data/experiment_data.json",
                       help="使用固定的实验数据文件（覆盖config.py中的设置）")
    parser.add_argument("--max-workers", type=int, default=1, 
                       help="单个模型内的最大并发数（默认使用config.py中的配置）")
    parser.add_argument("--model-workers", type=int, default=1, 
                       help="同时评估的模型数量（模型级别并发，默认使用config.py中的配置）")
    parser.add_argument("--models", type=str, nargs="+", default=None,
                       help="指定要评估的模型名称（必需，可指定多个），例如: --models Qwen3-8B")
    parser.add_argument("--max-history-tokens", type=int, default=None,
                       help="历史行为的最大Token数限制（覆盖数据集中的配置）")
    parser.add_argument("--max-history-days", type=int, default=None,
                       help="历史行为的最大天数限制（覆盖数据集中的配置）")
    parser.add_argument("--history-scene-filter", type=str, default=None,
                       help="历史行为场景过滤（覆盖数据集中的配置）")
    parser.add_argument("--history-process-mode", type=str, default="none",
                       choices=["none", "summary", "rag"],
                       help="历史行为处理模式: none(原始截断), summary(LangChain摘要), rag(LangChain RAG检索)")
    parser.add_argument("--USE_EXP_MEM", type=int) 
    parser.add_argument("--USE_EXP_MEM_PROGRESS", type=int, default=999) 
    parser.add_argument("--mem_name", type=str, default="experiment_data.json.practice") 
    args = parser.parse_args()
    


    print("=" * 50)
    print("参数配置如下：")
    print("=" * 50)
    for k, v in vars(args).items():
        print(f"{k:20s}: {v}")
    print("=" * 50)

    # LZQ # TODO: 一定要记得把prompt_builder中的ASR和OCR给恢复回来，在测试的时候
    global OUTPUT_DIR
    global RESULTS_DIR
    global USE_EXP_MEM
    global USE_EXP_MEM_PROGRESS
    global EXP_MEM_DIR_PATH 
    global EXP_MEM_DIR_PATH_PROGRESS 

    USE_EXP_MEM = args.USE_EXP_MEM
    USE_EXP_MEM_PROGRESS = args.USE_EXP_MEM_PROGRESS

    if USE_EXP_MEM == 0:
        EXP_MEM_DIR_PATH = "..*-*-*.."
        EXP_MEM_DIR_PATH_PROGRESS = "..*-*-*.."
        mem_name = args.use_fixed_data.split('/')[-1]
        if args.history_process_mode == "none":
            OUTPUT_DIR = f"./work_data/results/{mem_name}/{args.models[0]}_output_{args.max_history_tokens}"
            RESULTS_DIR = f"./work_data/results/{mem_name}/{args.models[0]}_results_{args.max_history_tokens}"
        else:
            OUTPUT_DIR = f"./work_data/results/{mem_name}/{args.models[0]}_output_{args.max_history_tokens}_{args.history_process_mode}"
            RESULTS_DIR = f"./work_data/results/{mem_name}/{args.models[0]}_results_{args.max_history_tokens}_{args.history_process_mode}"

    elif USE_EXP_MEM == 101:
        EXP_MEM_DIR_PATH = f"./work_data/exp_memory/{args.mem_name}"
        EXP_MEM_DIR_PATH_PROGRESS = f"./work_data/practice_progress/{args.mem_name}"
        if USE_EXP_MEM_PROGRESS == 999:
            OUTPUT_DIR = f"./work_data/results/{args.mem_name}/{args.models[0]}_output_{args.max_history_tokens}_101"
            RESULTS_DIR = f"./work_data/results/{args.mem_name}/{args.models[0]}_results_{args.max_history_tokens}_101"
        else:
            OUTPUT_DIR = f"./work_data/results/{args.mem_name}/{args.models[0]}_output_{args.max_history_tokens}_101_{USE_EXP_MEM_PROGRESS}"
            RESULTS_DIR = f"./work_data/results/{args.mem_name}/{args.models[0]}_results_{args.max_history_tokens}_101_{USE_EXP_MEM_PROGRESS}"

    else:
        raise ValueError(f"USE_EXP_MEM is {USE_EXP_MEM}")

    print("=" * 80)
    print("用户模拟评估系统 - 滚动预测评估")
    print("=" * 80)
    
    # 0. 配置历史处理模式（RAG/Summary）
    if args.history_process_mode != "none":
        print(f"\n📦 配置历史处理模式: {args.history_process_mode}")
        try:
            from history_processor import (
                set_history_process_mode, 
                _ensure_langchain_imported,
                get_history_processor
            )
            from config import RAG_CONFIG, SUMMARY_CONFIG
            
            # 调用检测函数，真正检测 LangChain 是否可用
            langchain_available = _ensure_langchain_imported()
            
            if not langchain_available:
                print("⚠️  警告: LangChain 未安装，历史处理模式将降级为 'none'")
                print("   安装命令: pip install langchain langchain-community langchain-openai faiss-cpu tiktoken")
            else:
                set_history_process_mode(args.history_process_mode)
                
                if args.history_process_mode == "rag":
                    embed_type = RAG_CONFIG.get('embedding_type', 'api')
                    embed_info = f"embedding_model={RAG_CONFIG['embedding_model']}"
                    if embed_type == "local":
                        embed_info += f", device={RAG_CONFIG.get('device', 'cuda')}"
                    print(f"   RAG 配置: top_k={RAG_CONFIG['top_k']}, type={embed_type}, {embed_info}")
                elif args.history_process_mode == "summary":
                    print(f"   Summary 配置: model={SUMMARY_CONFIG['model_name']}")
                
                # 【重要】在主线程预先初始化处理器，避免多线程竞争
                print("   正在预加载处理器（避免多线程冲突）...")
                processor = get_history_processor(args.history_process_mode)
                if processor is not None:
                    print("   ✓ 历史处理器预加载完成")
                else:
                    print("   ⚠️ 历史处理器预加载失败，将降级为普通模式")
        except ImportError as e:
            print(f"⚠️  警告: 无法导入 history_processor ({e})，历史处理模式将使用 'none'")
    
    # 1. 加载和准备数据
    print("\n步骤1: 加载和准备数据")
    print("-" * 80)
    
    # 确定数据文件路径
    fixed_data_path = args.use_fixed_data or FIXED_EXPERIMENT_DATA_PATH
    
    if not os.path.exists(fixed_data_path):
        print(f"❌ 错误：数据集不存在: {fixed_data_path}")
        print(f"请先运行: python prepare_experiment_data.py")
        return
    
    print(f"📌 使用实验数据集: {fixed_data_path}")
    eval_data, metadata = load_fixed_experiment_data(fixed_data_path)
    experiment_name = os.path.splitext(os.path.basename(fixed_data_path))[0]
    
    # 简化 experiment_name，去掉冗长的前缀
    # 例如: multi_scene_user_stats_filtered_output_top50_30t -> top50_30t
    redundant_prefixes = [
        "multi_scene_user_stats_filtered_output_",
        "multi_scene_user_stats_",
        "multi_scene_",
    ]
    for prefix in redundant_prefixes:
        if experiment_name.startswith(prefix):
            experiment_name = experiment_name[len(prefix):]
            break
    
    # 评估参数：优先使用命令行参数，否则使用 metadata 中的值
    max_history_tokens = args.max_history_tokens if args.max_history_tokens is not None else metadata.get('max_history_tokens')
    max_history_days = args.max_history_days if args.max_history_days is not None else metadata.get('max_history_days')
    history_scene_filter = args.history_scene_filter if args.history_scene_filter is not None else metadata.get('history_scene_filter')
    
    # 如果命令行参数覆盖了 metadata 中的值，打印提示
    if args.max_history_tokens is not None and args.max_history_tokens != metadata.get('max_history_tokens'):
        print(f"⚠️  命令行参数覆盖: max_history_tokens = {args.max_history_tokens} (数据集中为 {metadata.get('max_history_tokens')})")
    if args.max_history_days is not None and args.max_history_days != metadata.get('max_history_days'):
        print(f"⚠️  命令行参数覆盖: max_history_days = {args.max_history_days} (数据集中为 {metadata.get('max_history_days')})")
    if args.history_scene_filter is not None and args.history_scene_filter != metadata.get('history_scene_filter'):
        print(f"⚠️  命令行参数覆盖: history_scene_filter = {args.history_scene_filter} (数据集中为 {metadata.get('history_scene_filter')})")
    
    # 构建评估参数后缀（用于 results 文件名，与采样数据分离）
    # 这些参数不影响采样，只在评估时生效，需要在结果文件名中体现以区分不同配置
    eval_params_suffix = ""
    
    # 添加 max_history_tokens 到文件名
    if max_history_tokens and max_history_tokens > 0:
        # 转换为K单位，如 32000 -> 32k
        tokens_k = max_history_tokens // 1000
        if f"_{tokens_k}k" not in experiment_name and f"{tokens_k}k" not in experiment_name:
            eval_params_suffix += f"_{tokens_k}k"
        print(f"📊 历史行为最大Token数: {max_history_tokens} ({tokens_k}K)")
    
    # 添加 max_history_days 到文件名
    if max_history_days and max_history_days > 0:
        if f"_{max_history_days}d" not in experiment_name and f"{max_history_days}d" not in experiment_name:
            eval_params_suffix += f"_{max_history_days}d"
        print(f"📊 历史行为最大天数: {max_history_days} 天")
    
    # 添加 history_scene_filter 到文件名
    if history_scene_filter:
        # 场景名称简化映射
        scene_short_map = {
            "视频浏览": "video",
            "直播间": "live",
            "商城购物": "shop",
            "广告推荐": "ad",
            "电商客服对话": "chat"
        }
        scene_short = scene_short_map.get(history_scene_filter, history_scene_filter[:4])
        # 检查文件名是否已包含场景过滤信息
        if f"_hs{scene_short}" not in experiment_name and f"hs{scene_short}" not in experiment_name:
            eval_params_suffix += f"_hs{scene_short}"
        print(f"📊 历史场景过滤: {history_scene_filter} ({scene_short})")
    
    # 添加历史处理模式到文件名
    if args.history_process_mode != "none":
        mode_short = args.history_process_mode  # "summary" 或 "rag"
        if f"_{mode_short}" not in experiment_name:
            eval_params_suffix += f"_{mode_short}"
            if args.history_process_mode == "rag":
                from config import RAG_CONFIG
                eval_params_suffix += f"_k{RAG_CONFIG['top_k']}"
        print(f"📊 历史处理模式: {args.history_process_mode}")
    
    # 将评估参数后缀添加到 experiment_name
    if eval_params_suffix:
        experiment_name = f"{experiment_name}{eval_params_suffix}"
        print(f"📌 结果文件名将包含评估参数: {experiment_name}")
    
    if not eval_data:
        print("没有找到符合条件的用户数据，程序退出")
        return
    
    # 2. 确定要评估的模型
    print("\n步骤2: 确定评估模型")
    print("-" * 80)
    
    # 加载多模型配置
    try:
        from config import MODELS_TO_EVALUATE
        
        # 必须通过命令行参数指定要评估的模型
        if not args.models:
            print("❌ 错误：必须通过 --models 参数指定要评估的模型")
            print("\n可用的模型列表:")
            for m in MODELS_TO_EVALUATE:
                print(f"  - {m['name']}")
            print("\n使用示例:")
            print("  python evaluator.py --models Qwen2.5-72B-Instruct")
            print("  python evaluator.py --models Qwen2.5-72B-Instruct AWS-Claude-3.7")
            return
        
        # 根据指定的模型名称筛选
        models_to_run = [m for m in MODELS_TO_EVALUATE if m["name"] in args.models]
        
        # 检查是否有未找到的模型
        found_model_names = [m["name"] for m in models_to_run]
        not_found = [name for name in args.models if name not in found_model_names]
        
        if not_found:
            print(f"❌ 错误：以下模型在配置中未找到: {', '.join(not_found)}")
            print("\n可用的模型列表:")
            for m in MODELS_TO_EVALUATE:
                print(f"  - {m['name']}")
            return
        
        if not models_to_run:
            print("❌ 错误：没有找到要评估的模型")
            return
        
        print(f"✓ 找到 {len(models_to_run)} 个要评估的模型:")
        for m in models_to_run:
            model_info = m.get('model', 'AWS Bedrock' if m['type'] == 'aws_claude' else 'N/A')
            print(f"  - {m['name']} ({m['type']}, {model_info})")
    
    except ImportError:
        print("❌ 错误：config.py 中未找到 MODELS_TO_EVALUATE")
        print("请确保 config.py 中配置了 MODELS_TO_EVALUATE 列表")
        return
    
    # 2.5. 历史处理预计算（RAG 索引构建 或 Summary 摘要生成）
    if args.history_process_mode in ["rag", "summary"]:
        mode_name = "RAG 预嵌入" if args.history_process_mode == "rag" else "Summary 摘要生成"
        print(f"\n步骤2.5: {mode_name}（处理所有历史数据）")
        print("-" * 80)
        try:
            from history_processor import precompute_user_indices
            
            # 收集所有用户的所有历史行为统计
            all_actions = []
            base_total = 0
            test_total = 0
            empty_history_users = 0
            for idx, user_data in enumerate(eval_data):
                # 获取用户的基础历史行为（如9月份的行为）
                base_history = user_data.get("base_history") or []
                base_total += len(base_history)
                if base_history:
                    all_actions.extend(base_history)
                
                # 获取测试时间范围内的所有行为（如10-11月份）
                test_time_actions = user_data.get("test_time_all_actions") or []
                test_total += len(test_time_actions)
                if test_time_actions:
                    all_actions.extend(test_time_actions)

                if len(base_history) == 0 and len(test_time_actions) == 0:
                    empty_history_users += 1

                # 只打印前3个用户的样本统计，避免刷屏
                if idx < 3:
                    uid = user_data.get("user_id", "unknown")
                    print(f"  样本用户 {idx+1} ({uid}): base_history={len(base_history)}, test_time_all_actions={len(test_time_actions)}")
            
            # 去重（基于 action 内容）
            unique_actions = {}
            for action in all_actions:
                import hashlib
                import json
                key = hashlib.md5(json.dumps(action, sort_keys=True, ensure_ascii=False).encode()).hexdigest()
                if key not in unique_actions:
                    unique_actions[key] = action
            
            unique_list = list(unique_actions.values())
            print(f"📊 历史汇总: base_history={base_total}, test_time_all_actions={test_total}, 两者皆空的用户={empty_history_users}/{len(eval_data)}")
            print(f"📊 收集到 {len(all_actions)} 条历史行为（去重后 {len(unique_list)} 条）")

            # 按用户构建索引库或摘要
            if not eval_data:
                print("⚠️ eval_data 为空，跳过预处理")
            else:
                # precompute_user_indices 内部已有详细进度显示
                # 支持 rag 和 summary 两种模式
                # 对于 summary 模式，从配置中读取并发线程数
                max_workers = None
                if args.history_process_mode == "summary":
                    max_workers = SUMMARY_CONFIG.get("max_workers", None)
                    if max_workers:
                        print(f"   使用并发模式: {max_workers} 个线程")
                
                precompute_user_indices(eval_data, show_progress=True, mode=args.history_process_mode, max_workers=max_workers)
        except Exception as e:
            import traceback
            print(f"⚠️ 预处理失败: {e}")
            traceback.print_exc()
    
    # 3. 执行评估
    all_model_metrics = {}
    
    # 为本次运行生成统一的日期和时间戳（所有模型共享，使用上海时区）
    shanghai_tz = pytz.timezone('Asia/Shanghai')
    now_shanghai = datetime.now(shanghai_tz)
    run_date = now_shanghai.strftime("%Y-%m-%d")
    run_timestamp = now_shanghai.strftime("%Y-%m-%d_%H-%M-%S")
    
    print(f"\n本次运行时间（上海时区）: {run_timestamp}")
    print(f"所有模型结果将保存到:")
    print(f"  - output/{run_date}/{run_timestamp}/")
    print(f"  - results/{run_date}/{run_timestamp}/")
    
    # 确定模型级别的并发数
    model_level_workers = args.model_workers
    if model_level_workers is None:
        model_level_workers = 1  # 默认同时评估 1 个模型
    
    # 多模型模式
    print(f"\n步骤3: 执行多模型评估（共 {len(models_to_run)} 个模型）")
    print("=" * 80)
    print(f"模型级别并发数: {model_level_workers}")
    
    # 显示每个模型的并发配置
    if args.max_workers:
        print(f"单模型内并发数: {args.max_workers}（命令行指定，覆盖所有模型配置）")
    else:
        print("单模型内并发数: 使用各模型独立配置")
        for m in models_to_run:
            model_max_workers = m.get('max_workers', '未配置')
            print(f"  - {m['name']}: {model_max_workers}")
    print()
    
    def evaluate_single_model(model_config, idx, total, date, timestamp, max_history_tokens_config, max_history_days_config, history_scene_filter_config):
        """评估单个模型的函数"""
        model_name = model_config["name"]
        print(f"\n[{idx}/{total}] 开始评估模型: {model_name}")
        print("-" * 80)
        
        # 为每个模型创建独立的实验名称
        model_experiment_name = f"{experiment_name}_{model_name}"
        
        try:
            # 创建评估器（传入统一的日期和时间戳、max_history_tokens、max_history_days 和 history_scene_filter）
            evaluator = UserSimulationEvaluator(
                model_config=model_config,
                experiment_name=model_experiment_name,
                run_date=date,
                run_timestamp=timestamp,
                max_history_tokens=max_history_tokens_config,
                max_history_days=max_history_days_config,
                history_scene_filter=history_scene_filter_config
            )
            
            # 确定该模型的并发数
            # 优先级: 命令行参数 > 模型配置 > 全局配置
            model_workers = args.max_workers
            if model_workers is None:
                model_workers = model_config.get('max_workers', None)
            
            # 执行评估
            results = evaluator.evaluate_all(eval_data, max_workers=model_workers)
            
            # 计算指标
            print(f"\n计算 {model_name} 的评估指标...")
            metrics = evaluator.calculate_metrics(results)
            
            # 保存报告
            os.makedirs(RESULTS_DIR, exist_ok=True)
            report_path = f"{RESULTS_DIR}/{model_experiment_name}_evaluation_report.txt"
            evaluator.save_metrics_report(metrics, report_path)
            
            print(f"✓ {model_name} 评估完成")
            
            return model_name, metrics
            
        except Exception as e:
            print(f"❌ {model_name} 评估失败: {e}")
            import traceback
            traceback.print_exc()
            return model_name, None
    
    # 根据并发配置选择执行方式
    if model_level_workers == 1:
        # 串行执行
        print("💡 串行模式：按顺序评估每个模型\n")
        for idx, model_config in enumerate(models_to_run, 1):
            model_name, metrics = evaluate_single_model(model_config, idx, len(models_to_run), run_date, run_timestamp, max_history_tokens, max_history_days, history_scene_filter)
            if metrics:
                all_model_metrics[model_name] = metrics
    else:
        # 并发执行
        print(f"🚀 并发模式：最多同时评估 {model_level_workers} 个模型\n")
        
        with ThreadPoolExecutor(max_workers=model_level_workers) as executor:
            # 提交所有模型评估任务
            futures = []
            for idx, model_config in enumerate(models_to_run, 1):
                future = executor.submit(evaluate_single_model, model_config, idx, len(models_to_run), run_date, run_timestamp, max_history_tokens, max_history_days, history_scene_filter)
                futures.append(future)
            
            # 收集结果
            for future in as_completed(futures):
                try:
                    model_name, metrics = future.result()
                    if metrics:
                        all_model_metrics[model_name] = metrics
                except Exception as e:
                    print(f"❌ 模型评估任务异常: {e}")
                    import traceback
                    traceback.print_exc()
    
    # 生成对比报告
    if len(all_model_metrics) > 1:
        print("\n步骤4: 生成模型对比报告")
        print("-" * 80)
        comparison_path = f"{RESULTS_DIR}/{experiment_name}_model_comparison.txt"
        save_model_comparison_report(all_model_metrics, comparison_path, experiment_name, run_date, run_timestamp)
    
    # 打印摘要
    print("\n" + "=" * 80)
    print("评估完成！")
    print("=" * 80)
    
    for model_name, metrics in all_model_metrics.items():
        print(f"\n【{model_name}】")
        print(f"总预测次数: {metrics['overall']['total_predictions']}")
        print(f"API调用成功: {metrics['overall']['successful_api_calls']}")
        print(f"API成功率: {metrics['overall']['api_success_rate']:.2%}")
        print(f"解析成功率: {metrics['overall']['parsing_success_rate']:.2%}")
        
        if metrics["binary_metrics"]:
            print("\n二分类指标摘要:")
            for field, field_metrics in metrics["binary_metrics"].items():
                accuracy = field_metrics.get("Accuracy", float('nan'))
                auc = field_metrics.get("AUC", float('nan'))
                logloss = field_metrics.get("LogLoss", float('nan'))
                ece = field_metrics.get("ECE", float('nan'))
                
                print(f"  {field:20s}: Accuracy={accuracy*100:.2f}%, AUC={auc*100:.2f}%, LogLoss={logloss:.2f}, ECE={ece*100:.2f}%")
        
        if metrics["continuous_metrics"]:
            print("\n连续值指标摘要:")
            for field, field_metrics in metrics["continuous_metrics"].items():
                mae = field_metrics.get("MAE", float('nan'))
                rmse = field_metrics.get("RMSE", float('nan'))
                print(f"  {field:20s}: MAE={mae:.2f}, RMSE={rmse:.2f}")
        
        if metrics["text_metrics"]:
            print("\n文本指标摘要:")
            for field, field_metrics in metrics["text_metrics"].items():
                bleu = field_metrics.get("BLEU", float('nan'))
                char_f1 = field_metrics.get("CharF1", float('nan'))
                bert_f1 = field_metrics.get("BERTScore_F1", float('nan'))
                # 如果有 BERTScore，显示更详细的信息
                if bert_f1 and not (isinstance(bert_f1, float) and str(bert_f1) == 'nan'):
                    print(f"  {field:20s}: BLEU={bleu*100:.2f}%, CharF1={char_f1*100:.2f}%, BERTScore_F1={bert_f1*100:.2f}%")
                else:
                    print(f"  {field:20s}: BLEU={bleu*100:.2f}%, CharF1={char_f1*100:.2f}%")
    
    print("=" * 80)


def save_model_comparison_report(all_metrics: Dict, output_path: str, experiment_name: str, run_date: str = None, run_timestamp: str = None):
    """
    生成模型对比报告（按日期和时间分类，精确到秒，数值保留两位小数）
    
    Args:
        all_metrics: {model_name: metrics_dict}
        output_path: 输出文件路径
        experiment_name: 实验名称
        run_date: 运行日期（用于日期目录）
        run_timestamp: 运行时间戳（用于统一保存路径，精确到秒）
    """
    # 使用上海时区
    shanghai_tz = pytz.timezone('Asia/Shanghai')
    now_shanghai = datetime.now(shanghai_tz)
    
    # 使用统一的运行日期和时间戳
    date_dir = run_date or now_shanghai.strftime("%Y-%m-%d")
    timestamp = run_timestamp or now_shanghai.strftime("%Y-%m-%d_%H-%M-%S")
    
    # 修改输出路径，添加日期/时间两级子目录
    output_dir = os.path.dirname(output_path)
    filename = os.path.basename(output_path)
    timestamp_output_dir = os.path.join(output_dir, date_dir, timestamp)
    output_path = os.path.join(timestamp_output_dir, filename)
    
    os.makedirs(timestamp_output_dir, exist_ok=True)
    
    # 先将百分比指标转换为百分数形式（0.23 -> 23.0）
    all_metrics_percentage = {}
    for model_name, metrics in all_metrics.items():
        all_metrics_percentage[model_name] = convert_metrics_to_percentage(metrics)
    
    # 将浮点数保留两位小数
    all_metrics_rounded = round_floats(all_metrics_percentage, decimals=2)
    
    lines = []
    lines.append("=" * 100)
    lines.append("模型对比报告")
    lines.append(f"实验名称: {experiment_name}")
    lines.append("=" * 100)
    lines.append("")
    
    # 整体统计对比
    lines.append("## 整体统计对比")
    lines.append("-" * 100)
    lines.append(f"{'模型名称':<30s} {'总预测':<10s} {'API成功':<10s} {'API成功率':<12s} {'解析成功率':<12s}")
    lines.append("-" * 100)
    
    for model_name, metrics in all_metrics_rounded.items():
        overall = metrics["overall"]
        lines.append(
            f"{model_name:<30s} "
            f"{overall['total_predictions']:<10d} "
            f"{overall['successful_api_calls']:<10d} "
            f"{overall['api_success_rate']:<12.2f}% "
            f"{overall['parsing_success_rate']:<12.2f}%"
        )
    lines.append("")
    
    # 二分类指标对比
    if any(m.get("binary_metrics") for m in all_metrics_rounded.values()):
        lines.append("## 二分类指标对比")
        lines.append("-" * 100)
        
        # 收集所有字段
        all_fields = set()
        for metrics in all_metrics_rounded.values():
            all_fields.update(metrics.get("binary_metrics", {}).keys())
        
        for field in sorted(all_fields):
            lines.append(f"\n### {field}")
            lines.append(f"{'模型名称':<30s} {'Accuracy':<12s} {'AUC':<12s} {'LogLoss':<12s} {'ECE':<12s}")
            lines.append("-" * 100)
            
            for model_name, metrics in all_metrics_rounded.items():
                field_metrics = metrics.get("binary_metrics", {}).get(field, {})
                if field_metrics:
                    accuracy = field_metrics.get("Accuracy")
                    auc = field_metrics.get("AUC")
                    logloss = field_metrics.get("LogLoss")
                    ece = field_metrics.get("ECE")
                    
                    accuracy_str = f"{accuracy:.2f}" if accuracy is not None else "N/A"
                    auc_str = f"{auc:.2f}" if auc is not None else "N/A"
                    logloss_str = f"{logloss:.2f}" if logloss is not None else "N/A"
                    ece_str = f"{ece:.2f}" if ece is not None else "N/A"
                    
                    lines.append(
                        f"{model_name:<30s} "
                        f"{accuracy_str:<12s} "
                        f"{auc_str:<12s} "
                        f"{logloss_str:<12s} "
                        f"{ece_str:<12s}"
                    )
                else:
                    lines.append(f"{model_name:<30s} {'N/A':<12s} {'N/A':<12s} {'N/A':<12s} {'N/A':<12s}")
        lines.append("")
    
    # 连续值指标对比
    if any(m.get("continuous_metrics") for m in all_metrics_rounded.values()):
        lines.append("## 连续值指标对比")
        lines.append("-" * 100)
        
        # 收集所有字段
        all_fields = set()
        for metrics in all_metrics_rounded.values():
            all_fields.update(metrics.get("continuous_metrics", {}).keys())
        
        for field in sorted(all_fields):
            lines.append(f"\n### {field}")
            lines.append(f"{'模型名称':<30s} {'MAE':<12s} {'NMAE(%)':<12s} {'RMSE':<12s} {'R²':<12s}")
            lines.append("-" * 100)
            
            for model_name, metrics in all_metrics_rounded.items():
                field_metrics = metrics.get("continuous_metrics", {}).get(field, {})
                if field_metrics:
                    mae = field_metrics.get("MAE")
                    nmae = field_metrics.get("NMAE")
                    rmse = field_metrics.get("RMSE")
                    r2 = field_metrics.get("R²")
                    
                    mae_str = f"{mae:.2f}" if mae is not None else "N/A"
                    nmae_str = f"{nmae:.2f}" if nmae is not None else "N/A"
                    rmse_str = f"{rmse:.2f}" if rmse is not None else "N/A"
                    r2_str = f"{r2:.2f}" if r2 is not None else "N/A"
                    
                    lines.append(
                        f"{model_name:<30s} "
                        f"{mae_str:<12s} "
                        f"{nmae_str:<12s} "
                        f"{rmse_str:<12s} "
                        f"{r2_str:<12s}"
                    )
                else:
                    lines.append(f"{model_name:<30s} {'N/A':<12s} {'N/A':<12s} {'N/A':<12s} {'N/A':<12s}")
        lines.append("")
    
    lines.append("=" * 100)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("\n".join(lines))
    
    # 同时保存JSON格式（保留两位小数）
    json_path = output_path.replace(".txt", ".json")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump({
            "experiment_name": experiment_name,
            "models": all_metrics_rounded
        }, f, ensure_ascii=False, indent=2)
    
    print(f"✓ 模型对比报告已保存:")
    print(f"  - TXT:  {output_path}")
    print(f"  - JSON: {json_path}")


if __name__ == "__main__":
    main()
